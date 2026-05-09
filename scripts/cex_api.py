"""
cex_api.py — BLS Consumer Expenditure Survey (CES) Table 1300 wrapper.

CES Table 1300 — "Age of reference person" — publishes annual mean income,
income tax, and expenditure breakdowns across 7 age cohorts (Under 25,
25–34, 35–44, 45–54, 55–64, 65–74, 75+). Released each September for the
prior calendar year.

This module exposes:

  - fetch_table_1300_xlsx(year)  → download the published XLSX to cache
  - parse_table_1300(path)       → parse a cached XLSX into a normalized dict
  - cached_years()               → list years that have a cached file
  - SEEDED_LATEST_2023           → illustrative literal values used when no
                                   cached XLSX is present, so the dashboard
                                   can render a working visualization out
                                   of the box

Bot-blocking note: bls.gov sits behind Akamai and blocks programmatic file
downloads (HTTP 403). The download helper still attempts a fetch on the
chance of a future policy change, but the production refresh path is
manual: a human downloads the XLSX from
  https://www.bls.gov/cex/tables.htm
and drops it under data/context-cache/bls/cex/. The builder picks it up on
the next `python3 scripts/build-context.py` run and the seeded fallback
silently falls away.
"""

from __future__ import annotations

import sys
from pathlib import Path

# Reuse the existing project download helper. excel_scrape.download_to_cache
# already handles retries, custom UA, and idempotent caching.
import excel_scrape

HERE = Path(__file__).resolve().parent
PROJECT_ROOT = HERE.parent
CACHE_DIR = PROJECT_ROOT / "data" / "context-cache" / "bls" / "cex"

# Canonical BLS CEX URL pattern as of 2026-05. May change; if so, drop the
# file manually under CACHE_DIR.
TABLE_1300_URL_TEMPLATE = (
    "https://www.bls.gov/cex/tables/calendar-year/"
    "mean-item-share-average-standard-error/"
    "reference-person-age-ranges-{year}.xlsx"
)


# ---------------------------------------------------------------------------
# Download (best-effort) and cache discovery
# ---------------------------------------------------------------------------
def fetch_table_1300_xlsx(year: int, *, use_cache: bool = True) -> Path:
    """
    Best-effort fetch of CES Table 1300 for a given year. Returns the cache
    path when successful. Raises on HTTP failure (BLS commonly returns 403
    to non-browser User-Agents). Callers should wrap and continue.
    """
    filename = f"reference-person-age-ranges-{year}.xlsx"
    return excel_scrape.download_to_cache(
        url=TABLE_1300_URL_TEMPLATE.format(year=year),
        cache_subdir="bls/cex",
        filename=filename,
        use_cache=use_cache,
    )


def cached_years() -> list[int]:
    """Return sorted list of years for which a cached XLSX exists."""
    if not CACHE_DIR.exists():
        return []
    out: list[int] = []
    for p in CACHE_DIR.glob("reference-person-age-ranges-*.xlsx"):
        # Filename pattern: reference-person-age-ranges-YYYY.xlsx
        stem = p.stem
        try:
            year = int(stem.rsplit("-", 1)[-1])
        except ValueError:
            continue
        if 2000 <= year <= 2100:
            out.append(year)
    return sorted(out)


def cache_path_for_year(year: int) -> Path:
    return CACHE_DIR / f"reference-person-age-ranges-{year}.xlsx"


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------
# Cohort header strings as printed in BLS Table 1300 column headers, mapped
# to the internal CesAgeCohort identifier the wire format uses.
COHORT_HEADER_MAP: dict[str, str] = {
    "Under 25": "u25",
    "Under 25 years": "u25",
    "25-34": "a25_34",
    "25 to 34": "a25_34",
    "25-34 years": "a25_34",
    "35-44": "a35_44",
    "35 to 44": "a35_44",
    "35-44 years": "a35_44",
    "45-54": "a45_54",
    "45 to 54": "a45_54",
    "45-54 years": "a45_54",
    "55-64": "a55_64",
    "55 to 64": "a55_64",
    "55-64 years": "a55_64",
    "65-74": "a65_74",
    "65 to 74": "a65_74",
    "65-74 years": "a65_74",
    "75 and older": "a75plus",
    "75 years and older": "a75plus",
    "75+": "a75plus",
}


# ---------------------------------------------------------------------------
# Table 1300 layout: a single sheet "Table 1300", row 3 is the column header
# (Item, All consumer units, Under 25, 25-34, 35-44, 45-54, 55-64, 65 and
# older, 65-74, 75 and older). For every line item, col 1 holds the label
# (a section header with no values), and the next row's col 1 == "Mean" with
# the cohort values. Suppressed cells contain footnote markers (single
# letters "c"/"d"/"e") rather than values; we coerce those to None.
#
# Cohort column index → CesAgeCohort id (1-indexed openpyxl convention).
# We deliberately skip col 8 ("65 years and older") because it is the parent
# of cols 9 + 10, which we already split out.
# ---------------------------------------------------------------------------
COHORT_COLUMNS: list[tuple[int, str]] = [
    (3, "u25"),
    (4, "a25_34"),
    (5, "a35_44"),
    (6, "a45_54"),
    (7, "a55_64"),
    (9, "a65_74"),
    (10, "a75plus"),
]

# Known footnote / annotation suffixes that appear after some labels (e.g.
# "Federal income taxes b/"). Strip before matching.
FOOTNOTE_SUFFIXES = (" a/", " b/", " c/", " d/", " e/")


def _clean_label(s) -> str:
    if not isinstance(s, str):
        return ""
    out = s
    for suf in FOOTNOTE_SUFFIXES:
        out = out.split(suf)[0]
    return out.strip()


def _coerce_value(v) -> float | None:
    """Coerce a Mean-row cell into a float or None.

    BLS suppresses some values with single-letter footnote markers in place
    of numbers (typically 'c' for high RSE, 'd' for too-small, 'e' for no
    data). Anything non-numeric collapses to None.
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        # Treat 0.0 as suppressed only if BLS published a marker; here we
        # accept zeroes as real values (they appear for cohorts with no
        # measurable spending in some categories).
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s == "" or len(s) <= 2:
            # Single/double-letter footnote markers ('c', 'd', 'e').
            return None
        # Some files put commas in numbers stored as text.
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


# Acceptable label variants per concept. Order matters within a tuple — we
# pick the first matched row encountered scanning top-to-bottom (Table 1300
# repeats some items between the "Sources of income" block and earlier
# expenditure summary block; we want the income-block hit). For unique
# labels, a single-element tuple works.
LABEL_VARIANTS: dict[str, tuple[str, ...]] = {
    "wages":       ("Wages and salaries",),
    "selfEmp":     ("Self-employment income",),
    "socSec":      ("Social Security, private and government retirement",),
    "divRent":     (
        # 2023+ wording
        "Interest, dividends, rental income, and other property income",
        # 2014–2022 wording (missing "and")
        "Interest, dividends, rental income, other property income",
    ),
    "fedTax":      ("Federal income taxes",),
    "stateTax":    ("State and local income taxes",),
    # Spending category section headers (the Mean row is +1)
    "totalExp":    ("Average annual expenditures",),
    "food":        ("Food",),
    "housing":     ("Housing",),
    "transp":      ("Transportation",),
    "health":      ("Healthcare", "Health care"),
    "ent":         ("Entertainment",),
    "ins":         ("Personal insurance and pensions",),
}


def _find_mean_row(ws, label_variants: tuple[str, ...]) -> dict[str, float | None] | None:
    """Scan rows top-to-bottom for any of the label variants. When found,
    grab the next row (which should be 'Mean') and return cohort values."""
    target = set(label_variants)
    max_r = ws.max_row
    for r in range(1, max_r):
        cell = ws.cell(r, 1).value
        if _clean_label(cell) not in target:
            continue
        # Confirm the next row is the Mean row before reading values. Some
        # rare cases (rows under "Percent distribution:") use the same row
        # for label + value; those aren't in our target set anyway.
        next_label = ws.cell(r + 1, 1).value
        if _clean_label(next_label).lower() != "mean":
            # Not the structure we expect; skip and keep scanning.
            continue
        out: dict[str, float | None] = {}
        for col, cohort in COHORT_COLUMNS:
            out[cohort] = _coerce_value(ws.cell(r + 1, col).value)
        return out
    return None


def parse_table_1300(path: Path) -> dict | None:
    """
    Parse a cached CES Table 1300 XLSX into the on-wire shape used by the
    builder. Returns None if the file can't be parsed (schema drift,
    unexpected layout). Errors print to stderr; callers fall back to the
    seeded values for that year.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        print(
            "  cex parse: openpyxl required — `pip install openpyxl`",
            file=sys.stderr,
        )
        return None

    # Year is encoded in the filename: reference-person-age-ranges-{YYYY}.xlsx
    try:
        year = int(path.stem.rsplit("-", 1)[-1])
    except ValueError:
        print(f"  cex parse: {path.name}: cannot infer year from filename", file=sys.stderr)
        return None

    try:
        # NOTE: do NOT use read_only=True — that mode disables ws.cell()
        # random access, and the label-scan strategy below depends on it.
        # Files are <1MB so the extra memory is fine.
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"  cex parse: {path.name}: {type(e).__name__}: {e}", file=sys.stderr)
        return None

    # Use the "Table 1300" sheet if present; else the active sheet.
    ws = wb["Table 1300"] if "Table 1300" in wb.sheetnames else wb.active

    pulled: dict[str, dict[str, float | None] | None] = {}
    for key, variants in LABEL_VARIANTS.items():
        pulled[key] = _find_mean_row(ws, variants)

    # Verify the rows we hard-depend on were actually located. Allow a few
    # spending categories to fall through (the builder will treat None as a
    # row of None values), but if any of the income / tax / total-exp rows
    # are missing the year is treated as un-parseable.
    REQUIRED = ("wages", "selfEmp", "socSec", "divRent", "fedTax", "stateTax", "totalExp")
    missing = [k for k in REQUIRED if pulled.get(k) is None]
    if missing:
        print(
            f"  cex parse: {path.name}: missing required rows {missing}",
            file=sys.stderr,
        )
        return None

    # Compose the wire shape.
    def _sum_cohorts(a: dict[str, float | None] | None,
                     b: dict[str, float | None] | None) -> dict[str, float | None]:
        out: dict[str, float | None] = {}
        for _, cohort in COHORT_COLUMNS:
            av = a.get(cohort) if a else None
            bv = b.get(cohort) if b else None
            if av is None and bv is None:
                out[cohort] = None
            else:
                out[cohort] = (av or 0.0) + (bv or 0.0)
        return out

    def _empty_cohort_dict() -> dict[str, float | None]:
        return {cohort: None for _, cohort in COHORT_COLUMNS}

    def _diff_cohorts(total: dict[str, float | None],
                      *parts: dict[str, float | None] | None) -> dict[str, float | None]:
        out: dict[str, float | None] = {}
        for _, cohort in COHORT_COLUMNS:
            t = total.get(cohort)
            if t is None:
                out[cohort] = None
                continue
            running = t
            for p in parts:
                pv = p.get(cohort) if p else None
                running -= (pv or 0.0)
            out[cohort] = max(running, 0.0)
        return out

    # Round helper to keep wire format compact (whole dollars).
    def _round_dict(d: dict[str, float | None]) -> dict[str, float | None]:
        return {
            k: (None if v is None else int(round(v))) for k, v in d.items()
        }

    income = {
        "wagesBusiness":         _round_dict(_sum_cohorts(pulled["wages"], pulled["selfEmp"])),
        "socSecRetirement":      _round_dict(pulled["socSec"] or _empty_cohort_dict()),
        "dividendsInterestRent": _round_dict(pulled["divRent"] or _empty_cohort_dict()),
    }
    income_tax = {
        "federal":    _round_dict(pulled["fedTax"] or _empty_cohort_dict()),
        "stateLocal": _round_dict(pulled["stateTax"] or _empty_cohort_dict()),
    }

    spending_named = {
        "food":               pulled["food"] or _empty_cohort_dict(),
        "housing":            pulled["housing"] or _empty_cohort_dict(),
        "transportation":     pulled["transp"] or _empty_cohort_dict(),
        "healthcare":         pulled["health"] or _empty_cohort_dict(),
        "entertainment":      pulled["ent"] or _empty_cohort_dict(),
        "insurancePensions":  pulled["ins"] or _empty_cohort_dict(),
    }
    # All other = total annual expenditures − sum(named categories). Matches
    # the slide's framing; captures Apparel, Personal care, Reading,
    # Education, Tobacco, Misc, Cash contributions, Alcohol, etc.
    other = _diff_cohorts(
        pulled["totalExp"] or _empty_cohort_dict(),
        *spending_named.values(),
    )
    spending = {k: _round_dict(v) for k, v in spending_named.items()}
    spending["other"] = _round_dict(other)

    return {
        "year": year,
        "income": income,
        "incomeTax": income_tax,
        "spending": spending,
    }


# ---------------------------------------------------------------------------
# Seeded illustrative values (BLS Table 1300, 2023 calendar year)
# ---------------------------------------------------------------------------
# These are illustrative reference values used so the dashboard renders a
# working visualization before the real XLSX is dropped into the cache.
# Once a real XLSX is parsed, the builder uses parsed values instead and
# clears the `seeded` flag on the wire format.
#
# Cohort order in every dict: u25, a25_34, a35_44, a45_54, a55_64, a65_74,
# a75plus. Values are USD per consumer unit per year, rounded to nearest
# dollar, and reflect the slide pattern in NWCCOG's Economic Summit deck:
# income peaks 35–54, social security & retirement income dominates 65+,
# spending rotates from transportation/entertainment toward healthcare and
# housing as the reference-person cohort ages.
SEEDED_LATEST_2023: dict = {
    "year": 2023,
    "income": {
        # Wages + salaries + self-employment / business income
        "wagesBusiness": {
            "u25": 50620,
            "a25_34": 99480,
            "a35_44": 124560,
            "a45_54": 128300,
            "a55_64": 100440,
            "a65_74": 36870,
            "a75plus": 9320,
        },
        # Social Security, private + government retirement, pensions
        "socSecRetirement": {
            "u25": 540,
            "a25_34": 1620,
            "a35_44": 2480,
            "a45_54": 4980,
            "a55_64": 12380,
            "a65_74": 32560,
            "a75plus": 33580,
        },
        # Interest, dividends, rental, other property income
        "dividendsInterestRent": {
            "u25": 380,
            "a25_34": 1110,
            "a35_44": 2160,
            "a45_54": 3840,
            "a55_64": 5790,
            "a65_74": 6710,
            "a75plus": 5760,
        },
    },
    "incomeTax": {
        # Federal income taxes (positive USD paid)
        "federal": {
            "u25": 2640,
            "a25_34": 8960,
            "a35_44": 13560,
            "a45_54": 14820,
            "a55_64": 11240,
            "a65_74": 4180,
            "a75plus": 1820,
        },
        # State + local income taxes (positive USD paid)
        "stateLocal": {
            "u25": 940,
            "a25_34": 3210,
            "a35_44": 4720,
            "a45_54": 5180,
            "a55_64": 3960,
            "a65_74": 1480,
            "a75plus": 660,
        },
    },
    "spending": {
        "food": {
            "u25": 6080,
            "a25_34": 9240,
            "a35_44": 11460,
            "a45_54": 11020,
            "a55_64": 8970,
            "a65_74": 7610,
            "a75plus": 5680,
        },
        "housing": {
            "u25": 17220,
            "a25_34": 26140,
            "a35_44": 32200,
            "a45_54": 30310,
            "a55_64": 25890,
            "a65_74": 22340,
            "a75plus": 19580,
        },
        "transportation": {
            "u25": 8960,
            "a25_34": 13720,
            "a35_44": 16880,
            "a45_54": 16240,
            "a55_64": 12940,
            "a65_74": 9560,
            "a75plus": 5680,
        },
        "healthcare": {
            "u25": 1240,
            "a25_34": 3460,
            "a35_44": 4980,
            "a45_54": 6140,
            "a55_64": 7420,
            "a65_74": 8240,
            "a75plus": 8480,
        },
        "entertainment": {
            "u25": 1880,
            "a25_34": 3220,
            "a35_44": 4620,
            "a45_54": 4180,
            "a55_64": 3680,
            "a65_74": 3140,
            "a75plus": 1860,
        },
        "insurancePensions": {
            "u25": 3380,
            "a25_34": 8740,
            "a35_44": 11620,
            "a45_54": 12180,
            "a55_64": 9840,
            "a65_74": 3540,
            "a75plus": 1480,
        },
        "other": {
            "u25": 5680,
            "a25_34": 8420,
            "a35_44": 10840,
            "a45_54": 10520,
            "a55_64": 9280,
            "a65_74": 7780,
            "a75plus": 5860,
        },
    },
}


# Trend skeleton — a sparse set of historical years that frame the same
# story as the slide. Values are illustrative interpolations meant to
# convey the directional shape (gradual category shifts as cohorts age,
# nominal-dollar growth from 2014 → 2023). Replace by parsing real cached
# XLSX files when available.
SEEDED_TREND_YEARS: list[int] = [2014, 2017, 2020, 2023]


def _scale_cohort_dict(values: dict, factor: float) -> dict:
    return {k: round(v * factor) if v is not None else None for k, v in values.items()}


def build_seeded_trend() -> dict:
    """
    Build a trend dict matching the wire format from SEEDED_LATEST_2023.
    For each historical year, scale the 2023 latest values by an
    approximate nominal-USD deflator so the trend isn't a flat line.
    Real refresh: replace by parsing each cached XLSX year.
    """
    # Approximate nominal-USD scaling vs. 2023 (very rough; CPI-based).
    # 2014 ≈ 0.78, 2017 ≈ 0.85, 2020 ≈ 0.90, 2023 = 1.00.
    factors = {2014: 0.78, 2017: 0.85, 2020: 0.90, 2023: 1.00}

    paths: list[tuple[str, str]] = [
        ("income", "wagesBusiness"),
        ("income", "socSecRetirement"),
        ("income", "dividendsInterestRent"),
        ("incomeTax", "federal"),
        ("incomeTax", "stateLocal"),
        ("spending", "food"),
        ("spending", "housing"),
        ("spending", "transportation"),
        ("spending", "healthcare"),
        ("spending", "entertainment"),
        ("spending", "insurancePensions"),
        ("spending", "other"),
    ]

    trend: dict[str, list[dict]] = {}
    for group, cat in paths:
        key = f"{group}.{cat}"
        points: list[dict] = []
        latest_values = SEEDED_LATEST_2023[group][cat]
        for year in SEEDED_TREND_YEARS:
            scaled = _scale_cohort_dict(latest_values, factors[year])
            for cohort, val in scaled.items():
                points.append({"year": year, "cohort": cohort, "value": val})
        trend[key] = points
    return trend


if __name__ == "__main__":  # pragma: no cover
    print("cex_api — module entry point.", file=sys.stderr)
    print(f"Cache dir: {CACHE_DIR}", file=sys.stderr)
    print(f"Cached years: {cached_years()}", file=sys.stderr)
