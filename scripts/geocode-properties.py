#!/usr/bin/env python3
"""
geocode-properties.py — Resolve every unique Address that appears in the
Placer_resident_toplocations table to a (lat, lng) pair so build-placer.py
can ship property-level coordinates with the shopper JSON. Used by the
Activity map's Shoppers metric to render a genuine property-level heatmap
(MapLibre heatmap layer weighted by visits, same primitive as the
Workforce map heatmap).

Two-stage geocoder. Each address is normalized into 1–6 lookup variants
(strict → relaxed) and tried in order against OpenStreetMap Nominatim;
addresses Nominatim can't resolve are retried against the US Census
Geocoder (free, no API key, much more tolerant of unit/suite suffixes
and highway abbreviations). Successful matches are cached so subsequent
builds are near-instant.

Run:
    python3 scripts/geocode-properties.py        # incremental — only
                                                 # geocodes addresses not
                                                 # already in the cache.
    python3 scripts/geocode-properties.py --reset
                                                 # forces re-geocode of
                                                 # every address.
    python3 scripts/geocode-properties.py --retry-failed
                                                 # re-attempts addresses
                                                 # previously cached as
                                                 # failed or transient.
    python3 scripts/geocode-properties.py --no-census
                                                 # Nominatim-only run
                                                 # (no Census fallback).
    python3 scripts/geocode-properties.py --limit N
                                                 # stops after N new
                                                 # requests (useful for
                                                 # incremental fills /
                                                 # honoring Nominatim's
                                                 # bulk-use guidance).

Cache shape (data/placer/geocode-cache.json):
    {
      "address string": {"lat": 39.07, "lng": -108.54, "src": "nominatim", "queriedAt": "2026-05-16"},
      "address string": {"lat": 39.07, "lng": -108.54, "src": "census", "queriedAt": "...", "matchedAddress": "..."},
      "BAD_ADDRESS": {"failed": true, "queriedAt": "..."},      # negative cache (both providers said no)
      "FLAKY_ADDRESS": {"transient": true, "queriedAt": "..."}, # network error; --retry-failed will retry
      ...
    }

build-placer.py consumes the same JSON to attach coordinates to its
emitted shopper-property file.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
PROJECT_ROOT = HERE.parent
INPUT_PATH = PROJECT_ROOT / "data" / "placer" / "Placer.ai - Regional Activity .xlsx"
CACHE_PATH = PROJECT_ROOT / "data" / "placer" / "geocode-cache.json"
SHOPPERS_SHEET = "Retail Leakage"

USER_AGENT = (
    "glenwood-valley-behavioral-map/1.0 "
    "(transportation research; jake@glenwoodsprings.gov)"
)
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
CENSUS_URL = "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress"
# Nominatim usage policy: ≤1 request/sec, supply a real User-Agent.
NOMINATIM_SLEEP = 1.05
# Census's public endpoint is much more permissive; stay polite.
CENSUS_SLEEP = 0.5

# Project-area bounding box. Every address in the Placer workbook is a
# Colorado-corridor shopping destination, so a result outside this
# envelope is a Nominatim wrong-state mismatch (Carbondale CO vs.
# Carbondale IL is the canonical case). Buffer extends slightly into
# UT/NM/WY/NE/KS so legitimate edge geocodes survive.
BBOX_LAT_MIN, BBOX_LAT_MAX = 36.0, 42.0
BBOX_LNG_MIN, BBOX_LNG_MAX = -110.0, -101.0


def in_project_bbox(lat: float, lng: float) -> bool:
    return (
        BBOX_LAT_MIN <= lat <= BBOX_LAT_MAX
        and BBOX_LNG_MIN <= lng <= BBOX_LNG_MAX
    )


def load_cache(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        with path.open(encoding="utf-8") as fh:
            data = json.load(fh)
            if isinstance(data, dict):
                return data
    except json.JSONDecodeError:
        pass
    return {}


def save_cache(path: Path, cache: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp")
    with tmp.open("w", encoding="utf-8") as fh:
        json.dump(cache, fh, indent=2, sort_keys=True)
    tmp.replace(path)


def read_unique_addresses(xlsx_path: Path) -> list[str]:
    """Walk the Placer_resident_toplocations sheet (column E = address)
    and return the unique non-empty address strings in workbook order."""
    if not xlsx_path.exists():
        print(f"ERROR: workbook missing at {xlsx_path}", file=sys.stderr)
        sys.exit(1)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if SHOPPERS_SHEET not in wb.sheetnames:
        print(f"ERROR: sheet '{SHOPPERS_SHEET}' not found", file=sys.stderr)
        sys.exit(1)
    ws = wb[SHOPPERS_SHEET]
    seen: set[str] = set()
    ordered: list[str] = []
    for raw in ws.iter_rows(min_row=5, values_only=True):
        if not raw or len(raw) < 5:
            continue
        cell = raw[4]
        if not isinstance(cell, str):
            continue
        addr = cell.strip()
        if not addr:
            continue
        if addr in seen:
            continue
        seen.add(addr)
        ordered.append(addr)
    return ordered


_ZIP_RE = re.compile(r"\b(\d{5})(?:-\d{4})?\b")


def normalize_address(addr: str) -> str:
    """Trim + collapse whitespace so trivially-different variants hit the
    same cache entry."""
    return re.sub(r"\s+", " ", addr).strip()


# Unit/suite/apt/# tokens that Nominatim chokes on. Match a comma-prefixed
# form ("..., Unit 200") OR a bare trailing form (" Ste 18a, City, ...").
_UNIT_TOKEN = r"(?:Unit|Ste\.?|Suite|Apt\.?|Apartment|Bldg\.?|Building|Rm\.?|Room|#|No\.?)"
_UNIT_COMMA_RE = re.compile(
    rf",\s*{_UNIT_TOKEN}\s*[A-Za-z0-9\-/]+(?=\s*(?:,|$))",
    re.IGNORECASE,
)
_UNIT_BARE_RE = re.compile(
    rf"\s+{_UNIT_TOKEN}\s*[A-Za-z0-9\-/]+(?=\s*,)",
    re.IGNORECASE,
)
_LEADING_NUM_RE = re.compile(r"^\d+[A-Za-z]?\s+")


def _expand_highway_abbrevs(addr: str) -> str:
    """Replace OSM-unfriendly road-type abbreviations with their canonical
    long forms. Conservative — only touches the road component."""
    out = addr
    # I-70bl / I-70BL → I-70 Business Loop (Clifton/Grand Junction area)
    out = re.sub(r"\bI-70bl\b", "I-70 Business Loop", out, flags=re.IGNORECASE)
    # Generic <N>bl → <N> Business Loop (other I-/US- routes follow the
    # same convention).
    out = re.sub(r"\b(I-\d+|US-?\d+)bl\b", r"\1 Business Loop", out, flags=re.IGNORECASE)
    # US Hwy / Us Hwy → US Highway (Nominatim recognizes this form much
    # more reliably than the abbreviated one).
    out = re.sub(r"\bUS\s*Hwy\b", "US Highway", out, flags=re.IGNORECASE)
    # Co Rd / Co-Rd → County Road
    out = re.sub(r"\bCo\.?\s*Rd\b", "County Road", out, flags=re.IGNORECASE)
    # Rcr → Routt County Road (specific to NW Colorado addresses)
    out = re.sub(r"\bRcr\b", "Routt County Road", out, flags=re.IGNORECASE)
    # State highway prefixes — leave Highway 6/82/133/etc. canonical
    # "<N>" → "Highway <N>" only when Hwy appears bare without a route
    # designator.
    out = re.sub(r"\bHwy\b", "Highway", out, flags=re.IGNORECASE)
    # Co-82 / Co-133 → CO-82 / CO-133 (state highway uppercase prefix —
    # Nominatim is case-sensitive for this convention)
    out = re.sub(r"\bCo-(\d+)\b", r"CO-\1", out)
    return out


def _strip_units(addr: str) -> str:
    """Drop unit/suite/apt/# tokens in both comma-prefixed and bare
    trailing forms. Collapses any double-space left behind."""
    out = _UNIT_COMMA_RE.sub("", addr)
    out = _UNIT_BARE_RE.sub("", out)
    return re.sub(r"\s+", " ", out).strip()


def _fix_known_typos(addr: str) -> str:
    """Targeted substitutions for the handful of city-name typos
    appearing in the workbook. Keep this list tight — broad fuzzy
    matching at this stage causes wrong-state geocodes."""
    out = addr
    out = re.sub(r"\bGlenwood Spring\b(?!s)", "Glenwood Springs", out)
    out = re.sub(r"\bDebeque\b", "De Beque", out, flags=re.IGNORECASE)
    out = re.sub(r"\bTaugenbaugh\b", "Taughenbaugh", out, flags=re.IGNORECASE)
    return out


def geocoding_variants(addr: str) -> list[str]:
    """Return 1–6 deduplicated lookup variants in confidence order:
        1. raw normalized
        2. highway/road abbreviations expanded
        3. unit/suite/# stripped
        4. (2)+(3) combined
        5. known city-name typos fixed
        6. relaxed — leading street number dropped

    Variants are tried strict-first so a precise rooftop hit beats a
    coarser street-centroid hit when both are available."""
    seen: list[str] = []

    def push(v: str) -> None:
        v = normalize_address(v)
        if v and v not in seen:
            seen.append(v)

    raw = normalize_address(addr)
    push(raw)
    push(_expand_highway_abbrevs(raw))
    push(_strip_units(raw))
    push(_strip_units(_expand_highway_abbrevs(raw)))
    typo_fixed = _fix_known_typos(raw)
    if typo_fixed != raw:
        push(_strip_units(_expand_highway_abbrevs(typo_fixed)))
    # Relaxed fallback — drop the leading street number so Nominatim/Census
    # can resolve to the street centroid.
    relaxed = _LEADING_NUM_RE.sub("", _strip_units(_expand_highway_abbrevs(raw)))
    if relaxed and relaxed != raw:
        push(relaxed)
    return seen


def geocode_one_nominatim(address: str) -> dict | str | None:
    """One Nominatim lookup. Returns {lat, lng} on success, None on a
    confirmed-no-match, or the string 'transient' on a network error
    (caller surfaces this so a temporary failure doesn't poison the
    cache permanently)."""
    params = {
        "q": address,
        "format": "json",
        "limit": "1",
        "countrycodes": "us",
        "addressdetails": "0",
    }
    url = f"{NOMINATIM_URL}?{urllib.parse.urlencode(params)}"
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            payload = json.load(resp)
    except (urllib.error.URLError, json.JSONDecodeError, TimeoutError) as e:
        print(f"  nominatim network error: {e}", file=sys.stderr)
        return "transient"
    if not payload:
        return None
    first = payload[0]
    try:
        lat = float(first["lat"])
        lng = float(first["lon"])
    except (KeyError, ValueError, TypeError):
        return None
    if not in_project_bbox(lat, lng):
        return None
    return {"lat": round(lat, 6), "lng": round(lng, 6)}


def geocode_one_census(address: str) -> dict | str | None:
    """US Census Geocoder lookup — much more tolerant of unit/suite
    suffixes, highway abbreviations, and partial addresses than
    Nominatim. Returns {lat, lng, matchedAddress} on success, None on
    confirmed-no-match, 'transient' on network error."""
    params = {
        "address": address,
        "benchmark": "Public_AR_Current",
        "format": "json",
    }
    url = f"{CENSUS_URL}?{urllib.parse.urlencode(params)}"
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            payload = json.load(resp)
    except (urllib.error.URLError, json.JSONDecodeError, TimeoutError) as e:
        print(f"  census network error: {e}", file=sys.stderr)
        return "transient"
    try:
        matches = payload["result"]["addressMatches"]
    except (KeyError, TypeError):
        return None
    if not matches:
        return None
    first = matches[0]
    try:
        coords = first["coordinates"]
        # Census returns x = longitude, y = latitude
        lat = float(coords["y"])
        lng = float(coords["x"])
    except (KeyError, ValueError, TypeError):
        return None
    if not in_project_bbox(lat, lng):
        return None
    matched_addr = first.get("matchedAddress") or ""
    return {
        "lat": round(lat, 6),
        "lng": round(lng, 6),
        "matchedAddress": matched_addr,
    }


def is_pending_entry(entry: object) -> bool:
    """True when a cache value represents an unresolved (failed or
    transient-error) lookup that --retry-failed should re-attempt."""
    if not isinstance(entry, dict):
        return False
    return bool(entry.get("failed")) or bool(entry.get("transient"))


def main() -> int:
    ap = argparse.ArgumentParser(description="Geocode Placer property addresses.")
    ap.add_argument(
        "--reset",
        action="store_true",
        help="Discard existing cache and re-geocode every address.",
    )
    ap.add_argument(
        "--retry-failed",
        action="store_true",
        help="Re-attempt addresses cached as failed or transient.",
    )
    ap.add_argument(
        "--no-census",
        action="store_true",
        help="Disable the US Census Geocoder fallback (Nominatim only).",
    )
    ap.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Stop after N new geocode requests (default: no limit).",
    )
    args = ap.parse_args()

    addresses = read_unique_addresses(INPUT_PATH)
    print(f"Unique addresses in workbook: {len(addresses)}")

    cache = {} if args.reset else load_cache(CACHE_PATH)
    if cache:
        success = sum(
            1 for v in cache.values()
            if isinstance(v, dict) and not v.get("failed") and not v.get("transient")
        )
        failed = sum(1 for v in cache.values() if isinstance(v, dict) and v.get("failed"))
        transient = sum(1 for v in cache.values() if isinstance(v, dict) and v.get("transient"))
        print(
            f"Cache loaded: {len(cache)} entries "
            f"({success} hits, {failed} failed, {transient} transient)"
        )

    today = date.today().isoformat()
    new_requests = 0
    successes = 0
    successes_nominatim = 0
    successes_census = 0
    failures = 0
    transients = 0

    try:
        for i, addr in enumerate(addresses):
            norm = normalize_address(addr)
            cached = cache.get(norm)
            if cached is not None and not args.reset:
                # Skip resolved hits unconditionally; skip negative-cache
                # entries unless --retry-failed asks us to re-attempt.
                if not is_pending_entry(cached):
                    continue
                if not args.retry_failed:
                    continue
            if args.limit and new_requests >= args.limit:
                print(f"Hit --limit ({args.limit}); stopping for this run.")
                break
            new_requests += 1
            print(f"[{i + 1}/{len(addresses)}] {addr[:70]}")

            variants = geocoding_variants(addr)
            result: dict | None = None
            transient_seen = False

            # Pass 1 — Nominatim, strict variants first.
            for variant in variants:
                r = geocode_one_nominatim(variant)
                time.sleep(NOMINATIM_SLEEP)
                if r == "transient":
                    transient_seen = True
                    continue
                if isinstance(r, dict):
                    result = {**r, "src": "nominatim"}
                    if variant != norm:
                        result["matchedVariant"] = variant
                    successes_nominatim += 1
                    break

            # Pass 2 — Census fallback for whatever Nominatim couldn't
            # resolve (unless --no-census). Same variant ordering so a
            # cleaner string still beats a noisier one.
            if result is None and not args.no_census:
                for variant in variants:
                    r = geocode_one_census(variant)
                    time.sleep(CENSUS_SLEEP)
                    if r == "transient":
                        transient_seen = True
                        continue
                    if isinstance(r, dict):
                        result = {**r, "src": "census"}
                        if variant != norm:
                            result["matchedVariant"] = variant
                        successes_census += 1
                        break

            if result is not None:
                cache[norm] = {**result, "queriedAt": today}
                successes += 1
            elif transient_seen:
                # At least one provider returned a network error; flag as
                # transient so a future run can retry without --retry-failed
                # being needed to clear a "real" failure.
                cache[norm] = {"transient": True, "queriedAt": today}
                transients += 1
            else:
                cache[norm] = {"failed": True, "queriedAt": today}
                failures += 1

            # Persist every 25 requests so a Ctrl-C still saves progress.
            if new_requests % 25 == 0:
                save_cache(CACHE_PATH, cache)
    except KeyboardInterrupt:
        print("\nInterrupted — saving partial cache.")
    finally:
        save_cache(CACHE_PATH, cache)

    print(
        f"Done. New requests: {new_requests}. "
        f"Successes: {successes} ({successes_nominatim} nominatim, {successes_census} census). "
        f"Failures: {failures}. Transient: {transients}. "
        f"Cache: {CACHE_PATH.relative_to(PROJECT_ROOT)}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
