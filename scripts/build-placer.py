#!/usr/bin/env python3
"""
build-placer.py — Convert the Placer.ai Regional Activity workbook into five
parallel JSON files plus a summary index, all under ``public/data/placer/``.
Reads the workbook from ``data/placer/`` relative to the project root.
``data/placer/`` is gitignored so the proprietary workbook stays out of the
repo. Processes four origin-flow sheets (Employee/Visitor × Counts/Visits)
plus the ``Placer_resident_toplocations`` named table inside the Retail
Leakage sheet, which feeds the "Out of Market Shopping" metric.

Workbook layout (one sheet per metric):
    Row 2:  title cell ('Employee Origins by Zip Code - Employee Counts' etc)
    Row 4:  column headers: Destination Zip Code | Year | Origin Zipcode | City
            | State | lat | lng | % of <metric> | <metric value>
    Row 5+: data rows. Column A is always blank.

Output (one file per metric):
    {
      "metric": "employee-counts",
      "label": "Employee Origins — Employee Counts",
      "source": "Placer.ai",
      "lastBuilt": "<ISO-date>",
      "destAnchors": ["81601", "81623"],
      "paths": [["I70_E_GWS"], ["I70_W_GWS", "C82_GWS_CARB"], ...],
      "rows": [
        { "destZip": "81601", "originZip": "81601", "value": 5543,
          "pathId": 0 },
        ...
      ]
    }

Each row carries an integer ``pathId`` indexing the metric-level ``paths``
table, which holds every distinct corridor path used by rows in that metric.
This interning keeps file size in check: visitor-side metrics route most rows
through one of a handful of gateway-to-anchor paths, so a flat 73,000-row
list collapses to <200 unique paths. The path is computed by:
  1. Trying the LODES OD-pair lookup (preserved canonical paths from
     flows-inbound.json / flows-outbound.json).
  2. Else, resolving the origin to a known anchor node (zip→node from the
     corridor graph) or to GW_E / GW_W based on origin longitude (split at
     GATEWAY_SPLIT_LNG, matching build-data.py).
  3. Running Dijkstra over the corridor adjacency to the destination anchor.

This means Placer flows from outside the LODES Colorado-only universe (e.g.,
out-of-state visitor origins) still ride the corridor graph through the
appropriate I-70 gateway. The React side reads corridorPath off the row
directly — placerAdapters no longer needs to borrow from LODES on a miss.

Plus ``placer-summary.json`` with the same metadata plus per-metric row counts
so the React side can render shell metadata without parsing every file.

Run from any directory:
    python3 scripts/build-placer.py

Depends on:
  - public/data/corridors.json (emitted by build-data.py) for the corridor graph.
  - public/data/flows-inbound.json / flows-outbound.json (also from build-data.py)
    for the canonical LODES OD-pair corridor paths used as a first-choice lookup.
  Run build-data.py first if these don't exist.
"""

from __future__ import annotations

import heapq
import json
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
PROJECT_ROOT = HERE.parent
INPUT_PATH = PROJECT_ROOT / "data" / "placer" / "Placer.ai - Regional Activity .xlsx"
OUTPUT_DIR = PROJECT_ROOT / "public" / "data" / "placer"
CORRIDORS_JSON = PROJECT_ROOT / "public" / "data" / "corridors.json"
FLOWS_INBOUND_JSON = PROJECT_ROOT / "public" / "data" / "flows-inbound.json"
FLOWS_OUTBOUND_JSON = PROJECT_ROOT / "public" / "data" / "flows-outbound.json"
# Broader ZIP centroid lookup — used by the shopper metric to gateway-route
# destination ZIPs that aren't in the corridor graph (e.g., Grand Junction,
# Denver). Without this lookup the shopper flow's dest is "unresolved" and
# the corridor render dashes it out from the origin's centroid.
ZIPS_JSON = PROJECT_ROOT / "public" / "data" / "zips.json"
# Property-level geocode cache — populated by scripts/geocode-properties.py
# (Nominatim with caching). Used to attach a lat/lng to every shopper
# property so the Activity map's Shoppers heatmap can render at property
# granularity instead of falling back to dest-ZIP centroids.
GEOCODE_CACHE = PROJECT_ROOT / "data" / "placer" / "geocode-cache.json"

# Gateway routing constants — must match build-data.py so Placer flows enter
# through the same nodes as LODES flows.
GATEWAY_E_NODE = "GW_E"
GATEWAY_W_NODE = "GW_W"
GATEWAY_SPLIT_LNG = -107.3248

# Sheet name → (metric id, human label) mapping. Order matters only for the
# console output / summary structure.
SHEETS = [
    ("Employee Counts", "employee-counts", "Employee Origins — Employee Counts"),
    ("Employee Visits", "employee-visits", "Employee Origins — Employee Visits"),
    ("Visitor Counts",  "visitor-counts",  "Visitor Origins — Visitor Counts"),
    ("Visitor Visits",  "visitor-visits",  "Visitor Origins — Visitor Visits"),
]

# Shopper "Out of Market Shopping" metric is sourced from the named table
# `Placer_resident_toplocations` inside the Retail Leakage sheet. Schema is
# property-level (one row per resident-zip × destination-property) so the row
# layout differs from the four origin-flow sheets above — handled by its own
# reader (read_resident_toplocations) that extracts destination ZIPs from the
# Address column and aggregates Visits by (residentZip, destZip).
SHOPPERS_SHEET = "Retail Leakage"
SHOPPERS_TABLE = "Placer_resident_toplocations"
SHOPPERS_METRIC = "shoppers-toplocations"
SHOPPERS_LABEL = "Resident Top Locations — Out of Market Shopping"

# Address-column ZIP extractor. The address column ends in
# "..., CITY, STATE ZIP" for ~99.6% of rows, but the first 5-digit token in
# the string can also be a street number ("12345 Main St"). Use the
# end-preferred pattern first; fall back to the first-match pattern if the
# end-preferred pattern misses.
_ZIP_END_RE = re.compile(r"(\d{5})(?:-\d{4})?\b(?!.*\b\d{5}\b)")
_ZIP_ANY_RE = re.compile(r"\b(\d{5})(?:-\d{4})?\b")


def zip_str(value: object) -> str | None:
    """Coerce an Excel cell value into a 5-character zip code string. Returns
    None when the cell is blank or doesn't parse as a positive integer-ish
    number — those rows are skipped silently."""
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return None
        # Some cells already arrive as strings — keep as-is but pad.
        if v.isdigit():
            return v.zfill(5)
        return None
    if isinstance(value, (int, float)):
        n = int(value)
        if n <= 0:
            return None
        return f"{n:05d}"
    return None


def positive_int(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        n = int(value)
        return n if n > 0 else None
    if isinstance(value, str):
        v = value.strip()
        try:
            n = int(float(v))
            return n if n > 0 else None
        except ValueError:
            return None
    return None


def finite_float(value: object) -> float | None:
    """Coerce an Excel cell to a finite float, or None if absent/malformed.
    Used for origin lat/lng cells, which Placer always populates but might be
    surfaced as strings in some workbook exports."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        f = float(value)
        return f if (f == f and abs(f) < 1e9) else None  # rejects NaN/inf
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return None
        try:
            f = float(v)
            return f if f == f else None
        except ValueError:
            return None
    return None


def read_metric_sheet(ws) -> tuple[list[dict], list[str], int | None]:
    """Walk a metric sheet from row 5 onward. Returns
    (rows, dest_anchors, data_year).
    Column indexing (1-based, matches openpyxl):
        B (2) = Destination Zip Code
        C (3) = Year (data vintage — 2025 in the current workbook)
        D (4) = Origin Zipcode
        E (5) = City (origin city — kept on visitor rows so the client can
                      aggregate origin symbols by city instead of ZIP).
        F (6) = State (origin state — kept for visitor-row disambiguation).
        G (7) = lat
        H (8) = lng (used downstream for gateway classification)
        J (10) = numeric value
    The first row's Year is returned as data_year; the workbook is
    single-vintage so the value is consistent across rows in practice."""
    rows: list[dict] = []
    seen_dests: set[str] = set()
    data_year: int | None = None
    for raw in ws.iter_rows(min_row=5, values_only=True):
        if not raw or len(raw) < 10:
            continue
        dest = zip_str(raw[1])
        origin = zip_str(raw[3])
        city = raw[4] if isinstance(raw[4], str) else None
        state = raw[5] if isinstance(raw[5], str) else None
        lat = finite_float(raw[6])
        lng = finite_float(raw[7])
        value = positive_int(raw[9])
        if dest is None or origin is None or value is None:
            continue
        if data_year is None:
            y = positive_int(raw[2])
            if y is not None and 1900 < y < 3000:
                data_year = y
        rows.append({
            "destZip": dest,
            "originZip": origin,
            "value": value,
            "originLat": lat,
            "originLng": lng,
            "originCity": city.strip() if isinstance(city, str) else None,
            "originState": state.strip() if isinstance(state, str) else None,
        })
        seen_dests.add(dest)
    return rows, sorted(seen_dests), data_year


def extract_dest_zip(address: object) -> str | None:
    """Pull a 5-digit ZIP out of a property-address string. Prefer a ZIP at
    the END of the address ("..., CITY, STATE 81601") over a street number
    earlier in the string ("12345 Main St, Town, CO 81621"). Returns None if
    the address is missing or no 5-digit token is found."""
    if not isinstance(address, str):
        return None
    s = address.strip()
    if not s:
        return None
    m = _ZIP_END_RE.search(s)
    if m is not None:
        return m.group(1)
    # Fall back to first 5-digit token. Catches addresses where the ZIP isn't
    # at the very end (e.g., trailing whitespace, suite numbers after ZIP).
    m = _ZIP_ANY_RE.search(s)
    return m.group(1) if m is not None else None


def load_geocode_cache() -> dict[str, dict]:
    """Load address → {lat, lng} mappings from the Nominatim cache file
    written by scripts/geocode-properties.py. Returns an empty dict when
    the file is missing — callers fall back to the dest-ZIP centroid for
    properties that aren't in the cache."""
    if not GEOCODE_CACHE.exists():
        return {}
    try:
        with GEOCODE_CACHE.open(encoding="utf-8") as fh:
            data = json.load(fh)
        return data if isinstance(data, dict) else {}
    except json.JSONDecodeError:
        return {}


def read_resident_toplocations(ws) -> tuple[list[dict], list[str], int | None, int, list[dict]]:
    """Walk the Placer_resident_toplocations named table (B4:L11626) and
    aggregate Visits + Residents by (residentZip, destZip, groupCategory).
    The table covers one row per resident-zip × destination-property; this
    collapses to one row per resident-zip × destination-zip × category by
    summing Visits + Residents across all properties that share that key.
    Keeping the groupCategory dimension lets the client compute KPIs,
    category pie/rankings, and place rankings without re-aggregating.

    Filters out rows where destZip equals residentZip (in-market — same ZIP
    as home) so the emitted dataset is strictly out-of-market shopping
    flows.

    Column indexing (1-based, matches openpyxl):
        B (2) = Resident Zip Code (origin)
        C (3) = Year (data vintage)
        D (4) = Property (kept as a sample property name for tooltips)
        E (5) = Address (parsed for destination ZIP)
        F (6) = Sub Category (per-property venue type)
        G (7) = Residents (unique residents from this ZIP to this property)
        H (8) = Visits (aggregated)
        L (12) = Group Category (the rollup the client uses for pie + rank)

    Returns:
        (rows, sorted_origin_anchors, data_year, dropped_no_zip_count)
        - rows: list of {originZip, destZip, value, residents, category,
                          subCategory, propertySample, originLat: None,
                          originLng: None} dicts.
                originLat/Lng are None because the resident origin is always
                an anchor ZIP in the corridor graph — the lat/lng-based
                gateway routing isn't needed and would be wrong (we don't
                want to send a Glenwood resident through GW_W just because
                their resident ZIP lat happens to fall west of the split).
        - sorted_origin_anchors: unique resident ZIPs that appeared in the
                source data (the 11 valley ZIPs), used as the file's
                destAnchors field — these are the "focal anchors" for the
                shopper view, treated as origins rather than destinations.
        - data_year: first non-None Year cell encountered.
        - dropped_no_zip_count: how many rows had unparseable addresses.
                Logged to console for QA visibility on future workbook revs.
    """
    # key = (origin, dest, group_category) → aggregated buckets
    aggregated: dict[tuple[str, str, str], dict] = {}
    # key = address → property-level aggregate. Visits + residents sum
    # across every resident-zip that touches the same address so the
    # heatmap weight reflects total foot traffic per property.
    by_property: dict[str, dict] = {}
    origin_anchors: set[str] = set()
    data_year: int | None = None
    dropped_no_zip = 0
    geocodes = load_geocode_cache()
    geocoded_hits = 0
    geocoded_misses = 0
    for raw in ws.iter_rows(min_row=5, values_only=True):
        if not raw or len(raw) < 12:
            continue
        origin = zip_str(raw[1])
        property_name = raw[3] if isinstance(raw[3], str) else None
        address = raw[4]
        sub_category = raw[5] if isinstance(raw[5], str) else None
        residents = positive_int(raw[6]) or 0
        visits = positive_int(raw[7])
        group_category = raw[11] if isinstance(raw[11], str) else None
        if origin is None or visits is None:
            continue
        dest = extract_dest_zip(address)
        if dest is None:
            dropped_no_zip += 1
            continue
        if dest == origin:
            continue
        if data_year is None:
            y = positive_int(raw[2])
            if y is not None and 1900 < y < 3000:
                data_year = y
        cat = (group_category or "Other").strip()
        key = (origin, dest, cat)
        entry = aggregated.get(key)
        if entry is None:
            aggregated[key] = {
                "value": visits,
                "residents": residents,
                "subCategory": sub_category.strip() if isinstance(sub_category, str) else None,
                "propertySample": property_name.strip() if isinstance(property_name, str) else None,
            }
        else:
            entry["value"] += visits
            entry["residents"] += residents
            if entry["subCategory"] is None and isinstance(sub_category, str):
                entry["subCategory"] = sub_category.strip()
            if entry["propertySample"] is None and isinstance(property_name, str):
                entry["propertySample"] = property_name.strip()
        origin_anchors.add(origin)

        # Property-level rollup for the heatmap. Aggregate across all
        # resident ZIPs that visit the same address — the heatmap weight
        # is total foot traffic per property, not per (origin,property).
        # The per-anchor maps (`visitsByAnchor`, `residentsByAnchor`)
        # preserve the resident-ZIP breakdown so the client can re-rank
        # properties for a single selected anchor without re-aggregating
        # raw rows.
        if isinstance(address, str):
            addr_key = address.strip()
            if addr_key:
                prop_entry = by_property.get(addr_key)
                if prop_entry is None:
                    by_property[addr_key] = {
                        "address": addr_key,
                        "property": property_name.strip() if isinstance(property_name, str) else None,
                        "category": cat,
                        "subCategory": sub_category.strip() if isinstance(sub_category, str) else None,
                        "destZip": dest,
                        "visits": visits,
                        "residents": residents,
                        "visitsByAnchor": {origin: visits},
                        "residentsByAnchor": {origin: residents},
                    }
                else:
                    prop_entry["visits"] += visits
                    prop_entry["residents"] += residents
                    prop_entry["visitsByAnchor"][origin] = (
                        prop_entry["visitsByAnchor"].get(origin, 0) + visits
                    )
                    prop_entry["residentsByAnchor"][origin] = (
                        prop_entry["residentsByAnchor"].get(origin, 0) + residents
                    )

    # Attach geocode (if available) to each property entry.
    properties: list[dict] = []
    for addr, prop in by_property.items():
        geo = geocodes.get(addr)
        lat = None
        lng = None
        if isinstance(geo, dict) and not geo.get("failed"):
            lat = geo.get("lat")
            lng = geo.get("lng")
        if isinstance(lat, (int, float)) and isinstance(lng, (int, float)):
            geocoded_hits += 1
        else:
            geocoded_misses += 1
        properties.append({
            "address": prop["address"],
            "property": prop["property"],
            "category": prop["category"],
            "subCategory": prop["subCategory"],
            "destZip": prop["destZip"],
            "visits": prop["visits"],
            "residents": prop["residents"],
            "visitsByAnchor": prop["visitsByAnchor"],
            "residentsByAnchor": prop["residentsByAnchor"],
            "lat": round(float(lat), 6) if isinstance(lat, (int, float)) else None,
            "lng": round(float(lng), 6) if isinstance(lng, (int, float)) else None,
        })
    # Sort by visits descending so the heatmap renderer can cap at a
    # reasonable top-N without losing the densest properties.
    properties.sort(key=lambda p: p.get("visits", 0), reverse=True)
    print(
        f"      properties: {len(properties)} unique addresses "
        f"({geocoded_hits} geocoded, {geocoded_misses} pending — "
        f"run scripts/geocode-properties.py to fill the cache)"
    )

    rows: list[dict] = []
    for (origin, dest, cat), entry in aggregated.items():
        rows.append({
            "destZip": dest,
            "originZip": origin,
            "value": entry["value"],
            "residents": entry["residents"],
            "category": cat,
            "subCategory": entry["subCategory"],
            "propertySample": entry["propertySample"],
            "originLat": None,
            "originLng": None,
        })
    return rows, sorted(origin_anchors), data_year, dropped_no_zip, properties


# ---------------------------------------------------------------------------
# Corridor graph helpers (mirror build-data.py's logic so Placer flows ride
# the same nodes / gateways as LODES flows).
# ---------------------------------------------------------------------------
def load_corridor_graph() -> tuple[
    dict[str, dict],
    dict[str, list[tuple[str, str, float]]],
    dict[str, str],
]:
    """Read public/data/corridors.json (already built by build-data.py) and
    return (nodes, adjacency, zip_to_node). Uses the pre-routed lengthMeters
    rather than re-running OSRM."""
    if not CORRIDORS_JSON.exists():
        raise RuntimeError(
            f"corridors.json not found at {CORRIDORS_JSON}. Run "
            "`python3 scripts/build-data.py` first to build the corridor graph."
        )
    with CORRIDORS_JSON.open(encoding="utf-8") as fh:
        graph = json.load(fh)
    nodes: dict[str, dict] = {}
    zip_to_node: dict[str, str] = {}
    for n in graph.get("nodes", []):
        node_id = n.get("id")
        if not node_id:
            continue
        nodes[node_id] = n
        z = n.get("zip")
        if z:
            zip_to_node[z] = node_id
    adjacency: dict[str, list[tuple[str, str, float]]] = {nid: [] for nid in nodes}
    for c in graph.get("corridors", []):
        cid = c.get("id")
        a = c.get("from")
        b = c.get("to")
        length = float(c.get("lengthMeters") or 0.0)
        if not cid or a not in adjacency or b not in adjacency:
            continue
        adjacency[a].append((cid, b, length))
        adjacency[b].append((cid, a, length))
    return nodes, adjacency, zip_to_node


def shortest_corridor_path(
    adjacency: dict[str, list[tuple[str, str, float]]],
    start: str,
    end: str,
) -> list[str] | None:
    """Dijkstra over corridor-length-weighted adjacency. Ties on distance
    break on hop count, then on path tuple — matches build-data.py exactly so
    LODES and Placer agree on canonical paths."""
    if start == end:
        return []
    if start not in adjacency or end not in adjacency:
        return None
    pq: list[tuple[float, int, tuple[str, ...], str]] = [(0.0, 0, (), start)]
    best: dict[str, tuple[float, int, tuple[str, ...]]] = {start: (0.0, 0, ())}
    while pq:
        dist, hops, path, node = heapq.heappop(pq)
        if node == end:
            return list(path)
        prev = best.get(node)
        if prev is not None and (dist, hops, path) > prev:
            continue
        for cid, nbr, length in adjacency[node]:
            cand_dist = dist + length
            cand_hops = hops + 1
            cand_path = path + (cid,)
            existing = best.get(nbr)
            cand_key = (cand_dist, cand_hops, cand_path)
            if existing is None or cand_key < existing:
                best[nbr] = cand_key
                heapq.heappush(pq, (cand_dist, cand_hops, cand_path, nbr))
    return None


def gateway_for_lng(lng: float | None) -> str | None:
    """Classify an origin point as east- or west-gateway based on its
    longitude relative to GATEWAY_SPLIT_LNG. Unlike build-data.py's
    classify_external_zip, this is not restricted to 80/81 zip prefixes —
    Placer visitor origins span the whole country."""
    if lng is None:
        return None
    return GATEWAY_E_NODE if lng > GATEWAY_SPLIT_LNG else GATEWAY_W_NODE


def load_zip_centroids() -> dict[str, tuple[float, float]]:
    """Load (zip → (lat, lng)) from public/data/zips.json. Used by the
    shopper metric to gateway-route destination ZIPs that aren't part of
    the corridor graph but ARE in the broader ZipMeta universe (e.g.,
    Grand Junction 81504, Denver 80202). Returns an empty dict if the
    file is missing — callers fall back to dropping the row from the
    corridor render."""
    if not ZIPS_JSON.exists():
        return {}
    with ZIPS_JSON.open(encoding="utf-8") as fh:
        data = json.load(fh)
    if not isinstance(data, list):
        return {}
    out: dict[str, tuple[float, float]] = {}
    for entry in data:
        if not isinstance(entry, dict):
            continue
        z = entry.get("zip")
        lat = entry.get("lat")
        lng = entry.get("lng")
        if not isinstance(z, str):
            continue
        if not isinstance(lat, (int, float)) or not isinstance(lng, (int, float)):
            continue
        out[z] = (float(lat), float(lng))
    return out


def load_lodes_path_index() -> dict[tuple[str, str], list[str]]:
    """Build an (originZip, destZip) → corridorPath index from the already-
    emitted LODES flow files. Used as the first-choice lookup so Placer rows
    inherit the canonical LODES path on any OD pair LODES knows about."""
    index: dict[tuple[str, str], list[str]] = {}
    for path in (FLOWS_INBOUND_JSON, FLOWS_OUTBOUND_JSON):
        if not path.exists():
            continue
        with path.open(encoding="utf-8") as fh:
            data = json.load(fh)
        flows = data.get("flows") if isinstance(data, dict) else data
        if not isinstance(flows, list):
            continue
        for f in flows:
            ozip = f.get("originZip")
            dzip = f.get("destZip")
            cpath = f.get("corridorPath")
            if not isinstance(ozip, str) or not isinstance(dzip, str):
                continue
            if not isinstance(cpath, list):
                continue
            key = (ozip, dzip)
            # Inbound wins on collision (matches client flowsByOdKey precedence).
            if key not in index:
                index[key] = list(cpath)
    return index


def resolve_origin_node(
    origin_zip: str,
    origin_lng: float | None,
    zip_to_node: dict[str, str],
) -> str | None:
    """If the origin ZIP is bound to a corridor-graph node, use it. Otherwise
    classify by longitude into east or west gateway. Returns None only when
    both signals are absent."""
    if origin_zip in zip_to_node:
        return zip_to_node[origin_zip]
    return gateway_for_lng(origin_lng)


def attach_corridor_paths(
    rows: list[dict],
    adjacency: dict[str, list[tuple[str, str, float]]],
    zip_to_node: dict[str, str],
    lodes_paths: dict[tuple[str, str], list[str]],
    dest_zip_centroids: dict[str, tuple[float, float]] | None = None,
) -> dict[str, int]:
    """For each row, attach a `corridorPath` (list of corridor IDs). Mutates
    rows in place. Returns a counters dict for reporting.

    Resolution order per row:
      1. LODES OD-pair lookup — preserves canonical paths for any (origin,
         dest) pair the LODES build already routed.
      2. Self-loop (origin == dest) → empty path.
      3. Resolve origin to a node: known anchor ZIP → its node, else gateway
         by longitude.
      4. Resolve dest to a node: known anchor ZIP → its node. If the dest
         isn't in the corridor graph AND `dest_zip_centroids` is provided
         (shopper metric), gateway-classify by the dest ZIP's longitude
         so the path renders from the origin anchor out to the closest
         I-70 gateway instead of going unresolved.
      5. Dijkstra from origin node to dest node.

    Failures (no lng, dest not on the graph and no centroid lookup, no
    path) are surfaced as an empty path so the row remains in the stats
    but doesn't draw on the map. Each row also gets `originNode` for
    downstream diagnostics."""
    stats = {
        "lodes_hit": 0,
        "self_loop": 0,
        "routed_anchor": 0,
        "routed_gateway_e": 0,
        "routed_gateway_w": 0,
        "dest_gateway_e": 0,  # dest gateway-classified (shopper)
        "dest_gateway_w": 0,
        "unresolved": 0,
    }
    path_cache: dict[tuple[str, str], list[str] | None] = {}
    for row in rows:
        ozip = row["originZip"]
        dzip = row["destZip"]

        # 1. LODES OD-pair lookup wins so canonical paths stay in lockstep.
        lodes_path = lodes_paths.get((ozip, dzip))
        if lodes_path is not None:
            row["corridorPath"] = lodes_path
            row["originNode"] = zip_to_node.get(ozip)
            stats["lodes_hit"] += 1
            continue

        # 2. Self-loop short-circuit (matches build-data.py).
        if ozip == dzip:
            row["corridorPath"] = []
            row["originNode"] = zip_to_node.get(ozip)
            stats["self_loop"] += 1
            continue

        o_node = resolve_origin_node(ozip, row.get("originLng"), zip_to_node)
        d_node = zip_to_node.get(dzip)
        # 4. Shopper-style dest-fallback: out-of-graph destination → use
        # the dest ZIP's centroid (from zips.json) to gateway-classify and
        # route the origin anchor out to the closest I-70 gateway.
        if d_node is None and dest_zip_centroids is not None:
            centroid = dest_zip_centroids.get(dzip)
            if centroid is not None:
                _, dest_lng = centroid
                d_node = gateway_for_lng(dest_lng)
                if d_node == GATEWAY_E_NODE:
                    stats["dest_gateway_e"] += 1
                elif d_node == GATEWAY_W_NODE:
                    stats["dest_gateway_w"] += 1
        if o_node is None or d_node is None:
            row["corridorPath"] = []
            row["originNode"] = o_node
            stats["unresolved"] += 1
            continue

        cache_key = (o_node, d_node)
        if cache_key in path_cache:
            path = path_cache[cache_key]
        else:
            path = shortest_corridor_path(adjacency, o_node, d_node)
            path_cache[cache_key] = path
        row["corridorPath"] = path or []
        row["originNode"] = o_node

        if path is None:
            stats["unresolved"] += 1
        elif o_node == GATEWAY_E_NODE:
            stats["routed_gateway_e"] += 1
        elif o_node == GATEWAY_W_NODE:
            stats["routed_gateway_w"] += 1
        else:
            stats["routed_anchor"] += 1

    return stats


def main() -> int:
    if not INPUT_PATH.exists():
        print(f"ERROR: workbook not found at {INPUT_PATH}", file=sys.stderr)
        return 1

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Reading {INPUT_PATH.relative_to(PROJECT_ROOT)}…")
    nodes, adjacency, zip_to_node = load_corridor_graph()
    lodes_paths = load_lodes_path_index()
    print(
        f"  · corridor graph: {len(nodes)} nodes, "
        f"{sum(len(v) for v in adjacency.values()) // 2} edges"
    )
    print(f"  · LODES OD-pair path index: {len(lodes_paths)} pairs")

    wb = openpyxl.load_workbook(INPUT_PATH, data_only=True, read_only=True)
    iso_today = date.today().isoformat()

    # All four metric sheets share the same anchor set — capture the union as
    # we go and assert consistency for the user's sanity check.
    anchors_union: set[str] = set()
    summary_metrics: dict[str, dict] = {}
    # First non-None data_year across sheets wins. The workbook is
    # single-vintage in practice, so this is just defensive.
    data_year: int | None = None

    for sheet_name, metric_id, label in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  skip: sheet '{sheet_name}' not found", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        rows, dests, sheet_year = read_metric_sheet(ws)
        if data_year is None and sheet_year is not None:
            data_year = sheet_year
        anchors_union.update(dests)

        routing_stats = attach_corridor_paths(rows, adjacency, zip_to_node, lodes_paths)

        # Intern unique corridor paths so visitor metrics (where 70k+ rows
        # collapse to ~150 distinct paths) don't ship the same string arrays
        # 60,000 times. Order is stable: first-seen wins, so a re-run against
        # an unchanged workbook produces a byte-identical JSON.
        path_table: list[list[str]] = []
        path_id_by_key: dict[tuple[str, ...], int] = {}
        out_rows: list[dict] = []
        # Visitor metrics also retain origin lat/lng on each row so the
        # client can render origin symbols (e.g., a circle at each visitor
        # home ZIP centroid). Employee + shopper metrics drop the
        # coordinates — origin lookups via the local ZipMeta cover those
        # use cases. Rounded to 4 decimals (~10 m precision) to keep file
        # size in check.
        emit_origin_latlng = metric_id in ("visitor-counts", "visitor-visits")
        for r in rows:
            cpath = r["corridorPath"]
            key = tuple(cpath)
            pid = path_id_by_key.get(key)
            if pid is None:
                pid = len(path_table)
                path_id_by_key[key] = pid
                path_table.append(list(cpath))
            row_out: dict = {
                "destZip": r["destZip"],
                "originZip": r["originZip"],
                "value": r["value"],
                "pathId": pid,
            }
            if emit_origin_latlng:
                lat = r.get("originLat")
                lng = r.get("originLng")
                if isinstance(lat, (int, float)):
                    row_out["originLat"] = round(float(lat), 4)
                if isinstance(lng, (int, float)):
                    row_out["originLng"] = round(float(lng), 4)
                # City + state preserved for visitor metrics so the client
                # can aggregate origin symbols by City instead of ZIP — a
                # major city like Dallas spreads across 20+ ZIPs and the
                # ZIP-level bubbles fragment the visual story.
                city = r.get("originCity")
                state = r.get("originState")
                if isinstance(city, str) and city:
                    row_out["originCity"] = city
                if isinstance(state, str) and state:
                    row_out["originState"] = state
            out_rows.append(row_out)

        out_path = OUTPUT_DIR / f"placer-{metric_id}.json"
        payload = {
            "metric": metric_id,
            "label": label,
            "source": "Placer.ai",
            "lastBuilt": iso_today,
            "dataYear": sheet_year,
            "destAnchors": dests,
            "paths": path_table,
            "rows": out_rows,
        }
        out_path.write_text(json.dumps(payload, separators=(",", ":")) + "\n")
        print(
            f"  {metric_id}: {len(out_rows):>6} rows, {len(path_table):>3} unique paths "
            f"→ {out_path.relative_to(PROJECT_ROOT)}"
        )
        print(
            f"      routing: lodes={routing_stats['lodes_hit']}  "
            f"anchor={routing_stats['routed_anchor']}  "
            f"GW_E={routing_stats['routed_gateway_e']}  "
            f"GW_W={routing_stats['routed_gateway_w']}  "
            f"self={routing_stats['self_loop']}  "
            f"unresolved={routing_stats['unresolved']}"
        )
        summary_metrics[metric_id] = {
            "label": label,
            "rowCount": len(out_rows),
            "routing": routing_stats,
        }

    # Shopper metric: read the Placer_resident_toplocations named table from
    # the Retail Leakage sheet. Schema differs from the four origin sheets
    # (one row per resident-zip × property), so the reader aggregates Visits
    # by (residentZip, destZip) and parses dest ZIPs from the Address column.
    # Routing reuses attach_corridor_paths and the same path-interning code
    # path. destAnchors here lists ORIGINS (resident anchors) — for shoppers
    # the focal-anchor selection axis is flipped vs employee/visitor metrics.
    if SHOPPERS_SHEET in wb.sheetnames:
        # Note: workbook is opened read-only so `ws.tables` isn't available
        # to validate the named table's presence. The reader walks the sheet
        # by absolute cell columns (B / E / H starting row 5), which matches
        # the Placer_resident_toplocations table's layout. If the workbook
        # ever changes that layout the reader will surface zero usable rows.
        ws = wb[SHOPPERS_SHEET]
        rows, origin_anchors, sheet_year, dropped_no_zip, properties = (
            read_resident_toplocations(ws)
        )
        if not rows:
            print(
                f"  skip: '{SHOPPERS_SHEET}' yielded no shopper rows — "
                f"check column layout B (residentZip), E (address), "
                f"H (visits)",
                file=sys.stderr,
            )
        else:
            if data_year is None and sheet_year is not None:
                data_year = sheet_year

            # Provide a broader ZIP centroid lookup so shopper destinations
            # that aren't in the corridor graph (Grand Junction, Denver,
            # etc.) get gateway-classified by their longitude instead of
            # falling through as unresolved.
            zip_centroids = load_zip_centroids()
            routing_stats = attach_corridor_paths(
                rows, adjacency, zip_to_node, lodes_paths, zip_centroids
            )

            path_table: list[list[str]] = []
            path_id_by_key: dict[tuple[str, ...], int] = {}
            out_rows: list[dict] = []
            for r in rows:
                cpath = r["corridorPath"]
                key = tuple(cpath)
                pid = path_id_by_key.get(key)
                if pid is None:
                    pid = len(path_table)
                    path_id_by_key[key] = pid
                    path_table.append(list(cpath))
                row_out: dict = {
                    "destZip": r["destZip"],
                    "originZip": r["originZip"],
                    "value": r["value"],
                    "pathId": pid,
                    "residents": r.get("residents", 0),
                    "category": r.get("category") or "Other",
                }
                # Tooltip color — keep when present, omit when null to
                # keep the JSON tight.
                sub = r.get("subCategory")
                prop = r.get("propertySample")
                if isinstance(sub, str) and sub:
                    row_out["subCategory"] = sub
                if isinstance(prop, str) and prop:
                    row_out["propertySample"] = prop
                out_rows.append(row_out)

            out_path = OUTPUT_DIR / f"placer-{SHOPPERS_METRIC}.json"
            payload = {
                "metric": SHOPPERS_METRIC,
                "label": SHOPPERS_LABEL,
                "source": "Placer.ai",
                "lastBuilt": iso_today,
                "dataYear": sheet_year,
                "destAnchors": origin_anchors,
                "paths": path_table,
                "rows": out_rows,
                # Property-level points for the Shoppers Heatmap view.
                # Each entry carries a Nominatim-geocoded (lat, lng) when
                # available; properties without a cache hit have null
                # coords and the client falls back to dest-ZIP centroid.
                "properties": properties,
            }
            out_path.write_text(json.dumps(payload, separators=(",", ":")) + "\n")
            print(
                f"  {SHOPPERS_METRIC}: {len(out_rows):>6} rows, "
                f"{len(path_table):>3} unique paths "
                f"→ {out_path.relative_to(PROJECT_ROOT)}"
            )
            print(
                f"      origin anchors ({len(origin_anchors)}): {origin_anchors}"
            )
            print(
                f"      routing: lodes={routing_stats['lodes_hit']}  "
                f"anchor={routing_stats['routed_anchor']}  "
                f"GW_E={routing_stats['routed_gateway_e']}  "
                f"GW_W={routing_stats['routed_gateway_w']}  "
                f"destGW_E={routing_stats['dest_gateway_e']}  "
                f"destGW_W={routing_stats['dest_gateway_w']}  "
                f"self={routing_stats['self_loop']}  "
                f"unresolved={routing_stats['unresolved']}"
            )
            if dropped_no_zip:
                print(
                    f"      dropped {dropped_no_zip} row(s) with unparseable "
                    f"address (no 5-digit ZIP found)"
                )
            summary_metrics[SHOPPERS_METRIC] = {
                "label": SHOPPERS_LABEL,
                "rowCount": len(out_rows),
                "routing": routing_stats,
                "droppedNoZip": dropped_no_zip,
            }
    else:
        print(
            f"  skip: sheet '{SHOPPERS_SHEET}' not found — shopper metric "
            f"will be absent from output bundle",
            file=sys.stderr,
        )

    summary = {
        "lastBuilt": iso_today,
        "dataYear": data_year,
        "destAnchors": sorted(anchors_union),
        "metrics": summary_metrics,
    }
    summary_path = OUTPUT_DIR / "placer-summary.json"
    summary_path.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n")
    print(f"  summary: {summary_path.relative_to(PROJECT_ROOT)}")
    print(f"Done. Anchors covered: {sorted(anchors_union)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
