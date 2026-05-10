"""
ingest-nhgis-housing.py — Read the IPUMS NHGIS time-series Excel for
decennial housing units (1970 → 2020, reconciled to current geographic
boundaries) and emit normalized JSON for the housing builder.

Source file (manual drop):
    data/context-cache/sdo/Decennial Housing Counts.xlsx

Sheets used:
    nhgis0162_ts_nominal_place   — places, columns:
        (col 1) Year-Specific GIS Join Match Code
        (col 2) GEO_Yr (e.g., 'Glenwood Springs_1970')
        (col 3) Row Source Year
        (col 4) NHGIS Integrated State Name
        (col 5) NHGIS Integrated Place Name
        (col 6) Year-Specific Area Name
        (col 7) Decennial Housing units: Total

    nhgis0162_ts_nominal_county  — counties, columns:
        GISJOIN, YEAR, STATE, STATEFP, STATENH, COUNTY, COUNTYFP, COUNTYNH,
        NAME, A41AA (housing units total)

Filter:
- State = Colorado (NHGIS Integrated State Name == 'Colorado'; STATEFP == 8 for counties)
- Place rows: NHGIS Integrated Place Name matches one of our 11 anchor names
- County rows: COUNTYFP matches one of {37, 45, 77, 97} (Eagle, Garfield, Mesa, Pitkin)

Output (data/context-cache/sdo/nhgis-housing.json):
    {
      "places": {
        "0830780": [{"year": 1970, "value": 1574}, ..., {"year": 2020, "value": 4276}]
      },
      "counties": {
        "08045": [...],
        ...
      }
    }
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from geographies import STATE_FIPS, COUNTY_FIPS, PLACE_CODES

HERE = Path(__file__).resolve().parent
PROJECT_ROOT = HERE.parent
SRC_XLSX = PROJECT_ROOT / "data" / "context-cache" / "sdo" / "Decennial Housing Counts.xlsx"
OUT_JSON = PROJECT_ROOT / "data" / "context-cache" / "sdo" / "nhgis-housing.json"

# Place sheet header is on row 2 (1-indexed) — row 1 is blank.
# County sheet header is on row 1; row 2 is the human-readable label row.

PLACE_SHEET = "nhgis0162_ts_nominal_place"
COUNTY_SHEET = "nhgis0162_ts_nominal_county"


def _normalize_name(s: str) -> str:
    return (s or "").strip().lower()


def _build_name_to_geoid() -> dict[str, str]:
    out: dict[str, str] = {}
    for rec in PLACE_CODES.values():
        if rec.get("place_code") is None:
            continue
        out[_normalize_name(rec["place_name"])] = f"{STATE_FIPS}{rec['place_code']}"
    return out


def _to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        try:
            return int(float(v))
        except (ValueError, TypeError):
            return None


def main() -> int:
    if not SRC_XLSX.exists():
        print(f"ERROR: source not found: {SRC_XLSX}", file=sys.stderr)
        return 1
    try:
        import openpyxl  # type: ignore
    except ImportError:
        print("ERROR: openpyxl required (pip install openpyxl)", file=sys.stderr)
        return 1

    name_to_geoid = _build_name_to_geoid()

    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True, read_only=True)

    # --- Places ---------------------------------------------------------------
    ws_p = wb[PLACE_SHEET]
    p_rows_iter = ws_p.iter_rows(values_only=True)
    next(p_rows_iter)            # row 1 — blank
    p_header = next(p_rows_iter) # row 2 — column headers

    # Find columns we need by header text. Header has a leading None for col 0.
    p_idx = {}
    for i, c in enumerate(p_header):
        if c is None:
            continue
        p_idx[str(c).strip()] = i
    name_col = p_idx.get("NHGIS Integrated Place Name")
    state_col = p_idx.get("NHGIS Integrated State Name")
    year_col = p_idx.get("Row Source Year")
    units_col = p_idx.get("Decennial Housing units: Total")

    if None in (name_col, state_col, year_col, units_col):
        print(f"ERROR: missing place sheet columns. Got: {list(p_idx.keys())}", file=sys.stderr)
        return 1

    place_data: dict[str, dict[int, int | None]] = {g: {} for g in name_to_geoid.values()}
    p_rows_seen = 0
    p_rows_kept = 0
    for row in p_rows_iter:
        p_rows_seen += 1
        if (row[state_col] or "").strip() != "Colorado":
            continue
        nm = _normalize_name(str(row[name_col] or ""))
        geoid = name_to_geoid.get(nm)
        if not geoid:
            continue
        yr = _to_int(row[year_col])
        if yr is None or yr < 1970:
            continue
        units = _to_int(row[units_col])
        # Snowmass Village 1970 will be 0 / None pre-incorporation; preserve None
        # so the chart truncates rather than rendering a flat 0.
        place_data[geoid][yr] = units if units and units > 0 else None
        p_rows_kept += 1

    # --- Counties -------------------------------------------------------------
    ws_c = wb[COUNTY_SHEET]
    c_rows_iter = ws_c.iter_rows(values_only=True)
    c_header = next(c_rows_iter)          # row 1 — variable codes (GISJOIN, YEAR, STATEFP, COUNTYFP, A41AA)
    next(c_rows_iter)                     # row 2 — human labels (skip)

    c_idx = {str(c).strip(): i for i, c in enumerate(c_header) if c is not None}
    statefp_col = c_idx.get("STATEFP")
    countyfp_col = c_idx.get("COUNTYFP")
    cyear_col = c_idx.get("YEAR")
    cunits_col = c_idx.get("A41AA")

    if None in (statefp_col, countyfp_col, cyear_col, cunits_col):
        print(f"ERROR: missing county sheet columns. Got: {list(c_idx.keys())}", file=sys.stderr)
        return 1

    target_county_codes = {int(c) for c in COUNTY_FIPS.keys()}

    county_data: dict[str, dict[int, int | None]] = {
        f"{STATE_FIPS}{c}": {} for c in COUNTY_FIPS.keys()
    }
    c_rows_seen = 0
    c_rows_kept = 0
    for row in c_rows_iter:
        c_rows_seen += 1
        sfp = _to_int(row[statefp_col])
        if sfp != int(STATE_FIPS):
            continue
        cfp = _to_int(row[countyfp_col])
        if cfp not in target_county_codes:
            continue
        yr = _to_int(row[cyear_col])
        if yr is None or yr < 1970:
            continue
        units = _to_int(row[cunits_col])
        geoid = f"{STATE_FIPS}{str(cfp).zfill(3)}"
        county_data[geoid][yr] = units if units and units > 0 else None
        c_rows_kept += 1

    wb.close()

    out_places = {
        g: [{"year": y, "value": v} for y, v in sorted(d.items()) if v is not None]
        for g, d in place_data.items()
    }
    out_counties = {
        g: [{"year": y, "value": v} for y, v in sorted(d.items()) if v is not None]
        for g, d in county_data.items()
    }

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(
        json.dumps({"places": out_places, "counties": out_counties}, indent=2) + "\n"
    )

    print(f"Place rows scanned: {p_rows_seen}, anchor matches: {p_rows_kept}", file=sys.stderr)
    for g, s in out_places.items():
        years = [p["year"] for p in s]
        rng = f"{min(years)}–{max(years)}" if years else "no data"
        print(f"  place {g}: {len(s)} points ({rng})", file=sys.stderr)
    print(f"County rows scanned: {c_rows_seen}, anchor matches: {c_rows_kept}", file=sys.stderr)
    for g, s in out_counties.items():
        years = [p["year"] for p in s]
        rng = f"{min(years)}–{max(years)}" if years else "no data"
        print(f"  county {g}: {len(s)} points ({rng})", file=sys.stderr)
    print(f"Wrote {OUT_JSON.relative_to(PROJECT_ROOT)}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
