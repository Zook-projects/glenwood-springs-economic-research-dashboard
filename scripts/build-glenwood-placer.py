#!/usr/bin/env python3
"""build-glenwood-placer.py — Convert the four Glenwood-internal Placer.ai
workbooks into JSON consumed by the Activity Map's Glenwood scope.

Input workbooks (in ``data/placer/``, gitignored):
  - Tourist Visits Trends - Glenwood Springs.xlsx
  - Tourist Profile - Glenwood Springs.xlsx
  - Points of Interest.xlsx
  - Retail Hub - Visitation.xlsx

Output (under ``public/data/placer/glenwood/``):
  - visitation.json   — daily visits + annual metrics + avg daily visitors mix
  - retail-hubs.json  — per-hub daily visits, metrics, origin zips, profiles
  - pois.json         — per-POI monthly visits, metrics, origin zips, profiles
  - summary.json      — file-level index

Run from anywhere:
    python3 scripts/build-glenwood-placer.py
"""

from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from placer_helpers import finite_float, positive_int  # noqa: E402

HERE = Path(__file__).resolve().parent
PROJECT_ROOT = HERE.parent
INPUT_DIR = PROJECT_ROOT / "data" / "placer"
OUTPUT_DIR = PROJECT_ROOT / "public" / "data" / "placer" / "glenwood"

TRENDS_PATH = INPUT_DIR / "Tourist Visits Trends - Glenwood Springs.xlsx"
PROFILE_PATH = INPUT_DIR / "Tourist Profile - Glenwood Springs.xlsx"
POI_PATH = INPUT_DIR / "Points of Interest.xlsx"
RH_PATH = INPUT_DIR / "Retail Hub - Visitation.xlsx"

# Slug map. Display names from the workbooks → slugs that must match
# property.id in glenwood-features.geojson.
HUB_SLUGS = {
    "6th Street - East": "6th-street-east",
    "6th Street - West": "6th-street-west",
    "1400 Block Grand Ave": "1400-grand-ave",
    "Bethel & 7th Street": "bethel-7th",
    "Downtown Core": "downtown-core",
    "Glenwood Meadows": "glenwood-meadows",
    "Glenwood Springs Mall": "glenwood-springs-mall",
    "Roaring Fork Marketplace": "roaring-fork-marketplace",
}

POI_SLUGS = {
    "Hotel Overnight Stays": "hotel-overnight-stays",
    "Glenwood Caverns Adventure Park": "glenwood-caverns",
    "Hanging Lake": "hanging-lake",
    "Yampah Vapor Caves": "yampah-vapor-caves",
    "Glenwood Hot Springs Pool": "glenwood-hot-springs-pool",
    "Iron Mountain Hot Springs": "iron-mountain-hot-springs",
    "Two Rivers Park": "two-rivers-park",
    "Sunlight Mountain Resort": "sunlight-mountain-resort",
}


def iso_date(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return None
        # Best-effort parse — try a couple of common patterns.
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(v, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return None


def iso_month(value: object, month_name: object) -> str | None:
    """Combine a year cell and a month-name cell into 'YYYY-MM'."""
    if value is None or month_name is None:
        return None
    try:
        y = int(value)
    except (TypeError, ValueError):
        return None
    if not isinstance(month_name, str):
        return None
    months = {
        "january": 1, "february": 2, "march": 3, "april": 4, "may": 5,
        "june": 6, "july": 7, "august": 8, "september": 9, "october": 10,
        "november": 11, "december": 12,
    }
    m = months.get(month_name.strip().lower())
    if m is None:
        return None
    return f"{y:04d}-{m:02d}"


def positive_number(value: object) -> float | int | None:
    """Like positive_int but preserves floats (used for percentages / decimals)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        f = float(value)
        if f != f or abs(f) >= 1e12:
            return None
        if f.is_integer():
            return int(f)
        return f
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return None
        try:
            f = float(v)
            return int(f) if f.is_integer() else f
        except ValueError:
            return None
    return None


def parse_minutes(value: object) -> int | None:
    """Parse '180 min' or '1h 30m' or a bare number into total minutes."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        n = int(value)
        return n if n > 0 else None
    if isinstance(value, str):
        v = value.strip().lower()
        if not v:
            return None
        total = 0
        h = re.search(r"(\d+)\s*h", v)
        m = re.search(r"(\d+)\s*m", v)
        if h:
            total += int(h.group(1)) * 60
        if m:
            total += int(m.group(1))
        if total:
            return total
        # Maybe just a number string.
        try:
            return int(float(v))
        except ValueError:
            return None
    return None


# ---------------------------------------------------------------------------
# Visitation (Tourist Visits Trends + selected Tourist Profile fields)
# ---------------------------------------------------------------------------

DISTANCE_BANDS = {"0-25 mi", "25-50 mi", "50-100 mi", "100-250 mi", "250+ mi"}


def build_visitation() -> dict:
    wb = openpyxl.load_workbook(TRENDS_PATH, data_only=True, read_only=True)

    by_type: list[dict] = []
    by_distance: list[dict] = []
    by_overnight: list[dict] = []

    # Visits Combined NEW: header row 2, data row 3+
    # Columns: B Geography, C Date, D Visits, E Category, F Variable
    ws = wb["Visits Combined NEW"]
    min_date: str | None = None
    max_date: str | None = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 6:
            continue
        d = iso_date(row[2])
        visits = positive_int(row[3])
        category = row[4]
        variable = row[5]
        if d is None or visits is None or not category or not variable:
            continue
        if min_date is None or d < min_date:
            min_date = d
        if max_date is None or d > max_date:
            max_date = d
        if category == "Type":
            by_type.append({"date": d, "type": variable, "value": visits})
        elif category == "Distance":
            band = variable.strip()
            if band.startswith("0-25"):
                band_norm = "0-25"
            elif band.startswith("25-50"):
                band_norm = "25-50"
            elif band.startswith("50-100"):
                band_norm = "50-100"
            elif band.startswith("100-250"):
                band_norm = "100-250"
            elif band.startswith("250"):
                band_norm = "250+"
            else:
                continue
            by_distance.append({"date": d, "distance": band_norm, "value": visits})
        elif category == "Overnight":
            by_overnight.append({"date": d, "overnight": 1, "value": visits})

    # Visit Metrics: header row 2, data row 3+
    # Columns: B Geography, C Year, D Variable, E Metrics, F Value
    # Variable is a distance band ("0-25 mi" through "250+ mi") OR "Overnight".
    # Each (year, variable) combo carries the Metrics columns. We sum totals
    # over distance bands and compute a visits-weighted Avg. Days in Market
    # for the year. Overnight is a separate cross-cutting bucket that doesn't
    # contribute to the weighted distance average.
    ws = wb["Visit Metrics"]
    DISTANCE_BANDS = ("0-25 mi", "25-50 mi", "50-100 mi", "100-250 mi", "250+ mi")
    # year → { variable → { metric → value } }
    cells: dict[int, dict[str, dict[str, float]]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 6:
            continue
        try:
            year = int(row[2])
        except (TypeError, ValueError):
            continue
        variable = row[3]
        metric = row[4]
        val = row[5]
        if variable is None or metric is None or val is None:
            continue
        v_str = str(variable).strip()
        m_str = str(metric).strip()
        cells.setdefault(year, {}).setdefault(v_str, {})[m_str] = val

    annual_metrics: list[dict] = []
    days_by_distance_latest: dict[str, float] = {}
    latest_year_for_metrics: int | None = None
    for year in sorted(cells):
        per_var = cells[year]
        # Totals: sum visits and OOM visitors across distance bands.
        total_visits = 0
        total_oom = 0
        weighted_days_sum = 0.0
        weighted_days_weight = 0
        for band in DISTANCE_BANDS:
            band_metrics = per_var.get(band) or {}
            v = positive_int(band_metrics.get("Visits"))
            if v is not None:
                total_visits += v
            o = positive_int(band_metrics.get("Out-of-Market Visitors"))
            if o is not None:
                total_oom += o
            d = finite_float(band_metrics.get("Avg. Days in Market"))
            if d is not None and v is not None and v > 0:
                weighted_days_sum += v * d
                weighted_days_weight += v
        # First-row-wins style for daily-time metrics, anchored on 0-25 mi.
        first_band = per_var.get("0-25 mi") or {}
        avg_daily = first_band.get("Avg. Daily Time Spent in Market")
        median_daily = first_band.get("Median Daily Time Spent in Market")
        yoy = first_band.get("Visits YOY (1 Year Ago)")
        avg_daily_f = parse_minutes(avg_daily) if avg_daily is not None else None
        if avg_daily_f is None and avg_daily is not None:
            avg_daily_f = finite_float(avg_daily)
        median_daily_f = parse_minutes(median_daily) if median_daily is not None else None
        if median_daily_f is None and median_daily is not None:
            median_daily_f = finite_float(median_daily)
        yoy_f = finite_float(yoy)
        weighted_days = (
            round(weighted_days_sum / weighted_days_weight, 3)
            if weighted_days_weight > 0
            else None
        )
        annual_metrics.append(
            {
                "year": year,
                "visits": total_visits,
                "outOfMarketVisitors": total_oom,
                "avgDaysInMarket": weighted_days,
                "avgDailyTimeMinutes": int(avg_daily_f) if avg_daily_f is not None else None,
                "medianDailyTimeMinutes": int(median_daily_f) if median_daily_f is not None else None,
                "yoyPct": round(yoy_f, 4) if yoy_f is not None else None,
            }
        )
        latest_year_for_metrics = year

    # Per-distance Avg. Days in Market for the latest year — used by the
    # visitation KPI block when the user cross-filters to a specific band
    # (e.g., "0-25 mi" → 60.566 days). 'All' is the visits-weighted average.
    if latest_year_for_metrics is not None:
        per_var = cells[latest_year_for_metrics]
        total_visits_l = 0
        weighted_sum = 0.0
        for band in DISTANCE_BANDS:
            band_metrics = per_var.get(band) or {}
            d = finite_float(band_metrics.get("Avg. Days in Market"))
            v = positive_int(band_metrics.get("Visits"))
            if d is not None:
                days_by_distance_latest[band] = round(d, 3)
            if d is not None and v is not None and v > 0:
                weighted_sum += v * d
                total_visits_l += v
        overnight_metrics = per_var.get("Overnight") or {}
        d_over = finite_float(overnight_metrics.get("Avg. Days in Market"))
        if d_over is not None:
            days_by_distance_latest["Overnight"] = round(d_over, 3)
        if total_visits_l > 0:
            days_by_distance_latest["All"] = round(weighted_sum / total_visits_l, 3)

    # Average Daily Visitors: header row 2, data row 3+
    # Columns: B Type, C Daily Visitors
    avg_daily_mix: dict[str, int] = {}
    ws = wb["Average Daily Visitors"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 3:
            continue
        label = row[1]
        val = positive_int(row[2])
        if isinstance(label, str) and val is not None:
            avg_daily_mix[label.strip()] = val

    # Visitor demographic pull from Tourist Profile (only the "All" distance
    # subset, for the Visitation page's KPI block) plus per-distance Family
    # Households share (cross-filters to whichever distance band is active).
    visitor_profile: dict[str, object] = {}
    family_hh_by_distance: dict[str, float] = {}
    wb_profile = openpyxl.load_workbook(PROFILE_PATH, data_only=True, read_only=True)
    if "Tourist Profile" in wb_profile.sheetnames:
        ws_p = wb_profile["Tourist Profile"]
        # Header row 2: Geography | Census year | Geo_Yr | Time_range_year
        #               | Distance Index | Distance | Category | Key | Value
        #               | Percentage | National Index | Order
        all_keys = (
            ("Household Income", "Household Median Income", "_medianIncome"),
            ("Household Income", "Household Average Income", "_averageIncome"),
            ("Household Income", "Average Income per Person", "_avgIncomePerPerson"),
            ("Overview", "Population", "_population"),
        )
        wanted = {(c, k): out for c, k, out in all_keys}
        # Latest Time_range_year wins: track the max year seen for each
        # (distance, output-key) combo so the file ordering doesn't matter.
        latest_year_seen: dict[tuple[str, str], int] = {}
        latest_year_family: dict[str, int] = {}
        for row in ws_p.iter_rows(min_row=3, values_only=True):
            if len(row) < 10:
                continue
            try:
                time_range_year = int(row[3]) if row[3] is not None else None
            except (TypeError, ValueError):
                time_range_year = None
            distance = row[5]
            category = row[6]
            key = row[7]
            val = row[8]
            pct = row[9]
            if not category or not key:
                continue
            cat_s = str(category).strip()
            key_s = str(key).strip()
            tag = wanted.get((cat_s, key_s))
            if tag and distance == "All" and val is not None:
                prev_yr = latest_year_seen.get(("All", tag), -1)
                if time_range_year is None or time_range_year >= prev_yr:
                    v_val = positive_number(val)
                    if v_val is not None:
                        visitor_profile[tag] = (
                            round(float(v_val)) if tag != "_population" else v_val
                        )
                        latest_year_seen[("All", tag)] = (
                            time_range_year if time_range_year is not None else prev_yr
                        )
            # Family Households share, by distance (incl. "All").
            if cat_s == "Households" and key_s == "Family Households" and pct is not None:
                d_s = str(distance).strip() if distance else ""
                if d_s:
                    pct_f = finite_float(pct)
                    prev_yr = latest_year_family.get(d_s, -1)
                    if pct_f is not None and (
                        time_range_year is None or time_range_year >= prev_yr
                    ):
                        family_hh_by_distance[d_s] = round(float(pct_f), 5)
                        latest_year_family[d_s] = (
                            time_range_year if time_range_year is not None else prev_yr
                        )

    return {
        "source": "Placer.ai",
        "lastBuilt": date.today().isoformat(),
        "dataYearRange": (
            [int(min_date[:4]), int(max_date[:4])] if min_date and max_date else None
        ),
        "dataDateRange": [min_date, max_date],
        "dailyVisits": {
            "byType": by_type,
            "byDistance": by_distance,
            "byOvernight": by_overnight,
        },
        "annualMetrics": annual_metrics,
        "avgDailyVisitors": avg_daily_mix,
        "visitorProfile": visitor_profile,
        "daysInMarketByDistance": days_by_distance_latest,
        "familyHouseholdsPctByDistance": family_hh_by_distance,
    }


# ---------------------------------------------------------------------------
# Retail Hubs
# ---------------------------------------------------------------------------


def build_retail_hubs() -> dict:
    wb = openpyxl.load_workbook(RH_PATH, data_only=True, read_only=True)

    hubs: dict[str, dict] = {
        name: {
            "id": slug,
            "name": name,
            "dailyVisits": [],
            "metrics": {},
            "origins": [],
            "profile": {},
        }
        for name, slug in HUB_SLUGS.items()
    }

    # Retail Hubs - Visits: header row 2, data row 3+
    # Columns: B Geography, C Visitor Type, D Date, E Visits
    ws = wb["Retail Hubs - Visits"]
    min_date: str | None = None
    max_date: str | None = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 5:
            continue
        geo = row[1]
        visitor_type = row[2]
        d = iso_date(row[3])
        visits = positive_int(row[4])
        if not isinstance(geo, str) or geo not in hubs or d is None or visits is None:
            continue
        # Only aggregate the "All visitors" rows (skip Visitor Type breakouts
        # to avoid double-counting). Workbook uses different labels in some
        # rows — accept None / blank / "All Visitors" / "ALL".
        vt_norm = (visitor_type or "").strip().lower() if isinstance(visitor_type, str) else ""
        if vt_norm not in ("", "all visitors", "all"):
            continue
        hubs[geo]["dailyVisits"].append({"date": d, "value": visits})
        if min_date is None or d < min_date:
            min_date = d
        if max_date is None or d > max_date:
            max_date = d

    # Retail Hubs - Zip Visits: header row 2, data row 3+
    # Columns shifted right (Year starts at col D / index 3):
    #   D Year, E Geography, F Geo_Yr, G City State Year, H Zip Code, I City,
    #   J State, K lat, L lng, M % of Visits, N Visits, O YoY Change, P % YoY
    ws = wb["Retail Hubs - Zip Visits"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 16:
            continue
        try:
            year = int(row[3])
        except (TypeError, ValueError):
            continue
        geo = row[4]
        if not isinstance(geo, str) or geo not in hubs:
            continue
        zip_code = row[7]
        if zip_code is None:
            continue
        try:
            zip_str_val = f"{int(zip_code):05d}"
        except (TypeError, ValueError):
            continue
        lat = finite_float(row[10])
        lng = finite_float(row[11])
        pct = finite_float(row[12])
        visits = positive_int(row[13])
        yoy_pct = finite_float(row[15])
        if visits is None:
            continue
        hubs[geo]["origins"].append({
            "year": year,
            "zip": zip_str_val,
            "lat": round(lat, 4) if lat is not None else None,
            "lng": round(lng, 4) if lng is not None else None,
            "visits": visits,
            "pctOfVisits": round(pct, 6) if pct is not None else None,
            "yoyPct": round(yoy_pct, 4) if yoy_pct is not None else None,
        })

    # Extra Data Points: header row 2, data row 3+
    # Columns: B Geography, C Year, D Geo_Yr, E Attribute, F Value
    ws = wb["Extra Data Points"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 6:
            continue
        geo = row[1]
        year = row[2]
        attr = row[4]
        val = row[5]
        if not isinstance(geo, str) or geo not in hubs or year is None or attr is None:
            continue
        try:
            y_int = int(year)
        except (TypeError, ValueError):
            continue
        slot = hubs[geo]["metrics"].setdefault(
            str(y_int), {"avgDwellMin": None, "visitFrequency": None}
        )
        attr_s = str(attr).strip().lower()
        if "dwell" in attr_s:
            slot["avgDwellMin"] = parse_minutes(val)
        elif "frequency" in attr_s:
            f = finite_float(val)
            slot["visitFrequency"] = round(f, 3) if f is not None else None

    # Retail Hubs - Visitor Profile: header row 2, data row 3+
    # Columns: B Geography, C RH_Index, D Census year, E Geo_Yr, F Time_range_year,
    #          G Type, H Visitor Type, I Category, J Key, K Value,
    #          L Percentage, M National Index, N Order
    ws = wb["Retail Hubs - Visitor Profile"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 12:
            continue
        geo = row[1]
        visitor_type = row[7]
        category = row[8]
        key = row[9]
        val = row[10]
        pct = row[11]
        if not isinstance(geo, str) or geo not in hubs:
            continue
        # Only the "All Visitors" rows to keep payload simple.
        vt_norm = (visitor_type or "").strip().lower() if isinstance(visitor_type, str) else ""
        if vt_norm not in ("all visitors", "all", ""):
            continue
        if not category or not key:
            continue
        cat_s = str(category).strip()
        key_s = str(key).strip()
        prof = hubs[geo]["profile"]
        # Percentages stored under category buckets; raw values stashed under
        # sentinel keys for the KPI extractor (handled in the visitation flow
        # too).
        if pct is not None:
            p_val = finite_float(pct)
            if p_val is not None:
                prof.setdefault(cat_s, {})[key_s] = round(p_val, 6)
        else:
            v_val = positive_number(val)
            if v_val is not None:
                if cat_s == "Household Income" and key_s == "Household Median Income":
                    prof["_medianIncome"] = round(float(v_val))
                elif cat_s == "Household Income" and key_s == "Household Average Income":
                    prof["_averageIncome"] = round(float(v_val))
                elif cat_s == "Household Income" and key_s == "Average Income per Person":
                    prof["_avgIncomePerPerson"] = round(float(v_val))
                elif cat_s == "Overview" and key_s == "Population":
                    prof["_population"] = v_val

    return {
        "source": "Placer.ai",
        "lastBuilt": date.today().isoformat(),
        "dataYearRange": (
            [int(min_date[:4]), int(max_date[:4])] if min_date and max_date else None
        ),
        "hubs": list(hubs.values()),
    }


# ---------------------------------------------------------------------------
# Points of Interest
# ---------------------------------------------------------------------------


def build_pois() -> dict:
    wb = openpyxl.load_workbook(POI_PATH, data_only=True, read_only=True)

    pois: dict[str, dict] = {
        name: {
            "id": slug,
            "name": name,
            "monthlyVisits": [],
            "metrics": {},
            "origins": [],
            "originsLatLng": [],
            "profile": {},
        }
        for name, slug in POI_SLUGS.items()
    }

    # Zip Codes - Visits is the canonical source for both origin zips AND
    # monthly POI totals. The standalone "Visits" sheet is older — for the
    # latest months Placer only updates the Zip Codes - Visits sheet, and
    # within that sheet some months are POI-level summary rows (no zip)
    # where Placer didn't supply a zip-level breakdown. We capture both
    # row types here.
    #
    # Header row 2; data row 3+:
    #   B Year | C Month | D POI | E Geo_Yr | F Geo_Yr_Month |
    #   G City State Year | H Zip Code | I City | J State | K lat | L lng |
    #   M % of Visits | N Visits | O YoY Change | P % YoY
    ws = wb["Zip Codes - Visits"]
    latlng_acc: dict[tuple[str, str], dict] = {}
    # poi -> month -> visits (from POI-level summary rows, zip is None)
    poi_summary: dict[str, dict[str, int]] = {p: {} for p in pois}
    # poi -> month -> running zip-row sum (for months where no summary
    # row was supplied)
    poi_zip_sum: dict[str, dict[str, int]] = {p: {} for p in pois}

    min_month: str | None = None
    max_month: str | None = None

    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 16:
            continue
        try:
            year = int(row[1])
        except (TypeError, ValueError):
            continue
        month_name = row[2]
        poi = row[3]
        if not isinstance(poi, str) or poi not in pois:
            continue
        month = iso_month(year, month_name)
        if month is None:
            continue
        zip_code = row[7]
        visits = positive_int(row[13])
        if visits is None:
            continue

        if min_month is None or month < min_month:
            min_month = month
        if max_month is None or month > max_month:
            max_month = month

        if zip_code is None:
            # POI-level summary row: the only signal of monthly volume for
            # months where Placer didn't ship a zip breakdown (typical for
            # low-volume months at low-traffic POIs).
            poi_summary[poi][month] = visits
            continue

        try:
            zip_str_val = f"{int(zip_code):05d}"
        except (TypeError, ValueError):
            continue
        lat = finite_float(row[10])
        lng = finite_float(row[11])
        pct = finite_float(row[12])
        yoy_pct = finite_float(row[15])
        pois[poi]["origins"].append({
            "year": year,
            "month": month,
            "zip": zip_str_val,
            "visits": visits,
            "pctOfVisits": round(pct, 6) if pct is not None else None,
            "yoyPct": round(yoy_pct, 4) if yoy_pct is not None else None,
        })
        poi_zip_sum[poi][month] = poi_zip_sum[poi].get(month, 0) + visits
        # Accumulate origin-zip totals across all months for the map layer.
        key = (poi, zip_str_val)
        agg = latlng_acc.setdefault(
            key,
            {
                "zip": zip_str_val,
                "lat": round(lat, 4) if lat is not None else None,
                "lng": round(lng, 4) if lng is not None else None,
                "totalVisits": 0,
            },
        )
        agg["totalVisits"] += visits
        # Backfill coordinates if a later row had them and an earlier didn't.
        if agg["lat"] is None and lat is not None:
            agg["lat"] = round(lat, 4)
        if agg["lng"] is None and lng is not None:
            agg["lng"] = round(lng, 4)

    for (poi, _zip), entry in latlng_acc.items():
        if entry["lat"] is None or entry["lng"] is None:
            continue
        pois[poi]["originsLatLng"].append(entry)

    # Build monthlyVisits per POI by merging the POI-level summary rows
    # with the per-month sum of zip-level rows. POI-level summary wins
    # when both are present (they agree where both exist; the summary is
    # the authoritative number when zip rows are sparse).
    for poi in pois:
        merged: dict[str, int] = dict(poi_zip_sum[poi])
        for month, visits in poi_summary[poi].items():
            merged[month] = visits
        pois[poi]["monthlyVisits"] = [
            {"date": m, "value": v}
            for m, v in sorted(merged.items())
        ]

    # Extra Data: header row 2, data row 3+
    # Columns: B Geography, C Year, D Geo_Yr, E Attribute, F Value
    ws = wb["Extra Data"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 6:
            continue
        geo = row[1]
        year = row[2]
        attr = row[4]
        val = row[5]
        if not isinstance(geo, str) or geo not in pois or year is None or attr is None:
            continue
        try:
            y_int = int(year)
        except (TypeError, ValueError):
            continue
        slot = pois[geo]["metrics"].setdefault(
            str(y_int), {"avgDwellMin": None, "visitFrequency": None}
        )
        attr_s = str(attr).strip().lower()
        if "dwell" in attr_s:
            slot["avgDwellMin"] = parse_minutes(val)
        elif "frequency" in attr_s:
            f = finite_float(val)
            slot["visitFrequency"] = round(f, 3) if f is not None else None

    # Profiles: header row 2, data row 3+
    # Columns: A Census year, B Time_range_year, C POI Index, D POI,
    #          E Category, F Key, G Value, H Percentage, I National Index, J Order
    ws = wb["Profiles"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 9:
            continue
        poi = row[3]
        category = row[4]
        key = row[5]
        val = row[6]
        pct = row[7]
        if not isinstance(poi, str) or poi not in pois:
            continue
        if not category or not key:
            continue
        cat_s = str(category).strip()
        key_s = str(key).strip()
        prof = pois[poi]["profile"]
        if pct is not None:
            p_val = finite_float(pct)
            if p_val is not None:
                prof.setdefault(cat_s, {})[key_s] = round(p_val, 6)
        else:
            v_val = positive_number(val)
            if v_val is not None:
                if cat_s == "Household Income" and key_s == "Household Median Income":
                    prof["_medianIncome"] = round(float(v_val))
                elif cat_s == "Household Income" and key_s == "Household Average Income":
                    prof["_averageIncome"] = round(float(v_val))
                elif cat_s == "Household Income" and key_s == "Average Income per Person":
                    prof["_avgIncomePerPerson"] = round(float(v_val))
                elif cat_s == "Overview" and key_s == "Population":
                    prof["_population"] = v_val

    return {
        "source": "Placer.ai",
        "lastBuilt": date.today().isoformat(),
        "dataMonthRange": [min_month, max_month],
        "pois": list(pois.values()),
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> int:
    missing = [p for p in (TRENDS_PATH, PROFILE_PATH, POI_PATH, RH_PATH) if not p.exists()]
    if missing:
        for p in missing:
            print(f"ERROR: missing workbook {p}", file=sys.stderr)
        return 1

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Building Glenwood Placer.ai bundle…")

    print(f"  · visitation  ← {TRENDS_PATH.name}")
    visitation = build_visitation()
    (OUTPUT_DIR / "visitation.json").write_text(
        json.dumps(visitation, separators=(",", ":")) + "\n"
    )
    print(
        f"      daily rows: byType={len(visitation['dailyVisits']['byType'])} "
        f"byDistance={len(visitation['dailyVisits']['byDistance'])} "
        f"byOvernight={len(visitation['dailyVisits']['byOvernight'])} "
        f"annual={len(visitation['annualMetrics'])}"
    )

    print(f"  · retail-hubs ← {RH_PATH.name}")
    retail_hubs = build_retail_hubs()
    (OUTPUT_DIR / "retail-hubs.json").write_text(
        json.dumps(retail_hubs, separators=(",", ":")) + "\n"
    )
    for h in retail_hubs["hubs"]:
        print(
            f"      {h['id']:>30}  visits={len(h['dailyVisits']):>5}  "
            f"origins={len(h['origins']):>5}  metrics={list(h['metrics'].keys())}"
        )

    print(f"  · pois        ← {POI_PATH.name}")
    pois = build_pois()
    (OUTPUT_DIR / "pois.json").write_text(
        json.dumps(pois, separators=(",", ":")) + "\n"
    )
    for p in pois["pois"]:
        print(
            f"      {p['id']:>30}  monthly={len(p['monthlyVisits']):>4}  "
            f"origins={len(p['origins']):>6}  latlng={len(p['originsLatLng']):>5}"
        )

    summary = {
        "source": "Placer.ai",
        "lastBuilt": date.today().isoformat(),
        "files": {
            "visitation": {
                "dataDateRange": visitation["dataDateRange"],
                "annualYears": [m["year"] for m in visitation["annualMetrics"]],
            },
            "retailHubs": {
                "hubCount": len(retail_hubs["hubs"]),
                "totalDailyRows": sum(len(h["dailyVisits"]) for h in retail_hubs["hubs"]),
            },
            "pois": {
                "poiCount": len(pois["pois"]),
                "totalOriginRows": sum(len(p["origins"]) for p in pois["pois"]),
            },
        },
    }
    (OUTPUT_DIR / "summary.json").write_text(
        json.dumps(summary, separators=(",", ":")) + "\n"
    )

    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
